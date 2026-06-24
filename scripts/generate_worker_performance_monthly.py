import json
import math
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "raw" / "employee-performance-system"
RAW_DATA_LABEL = "data/raw/employee-performance-system"
OUTPUT_JSON = ROOT / "data" / "worker-performance-monthly.json"
OUTPUT_JS = ROOT / "data" / "worker-performance-monthly.js"
TOOLS_ROSTER_FILE = "tools_101_employee_roster.xlsx"

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_ORDER = (
    [f"{year} {month}" for year in range(2022, 2026) for month in MONTH_NAMES]
    + [f"2026 {month}" for month in MONTH_NAMES[:5]]
)
EXPORT_MONTH_COLUMNS = ["last Dec", "last year aver.", *MONTH_NAMES]
SUPPLEMENTAL_TRANSFER_REASONS = {"noEmployeeMonthRecord", "unconfirmedRequester", "unconfirmedResponsible"}
MTTR_DAILY_ERROR_THRESHOLD_MINUTES = 1440
MTTR_MONTHLY_ERROR_THRESHOLD_MINUTES = 1440
REPAIR_HOURS_QUALITY_PERCENTILE = 0.99
REPAIR_HOURS_QUALITY_REASON = "repairHoursExceedsP99"
REPAIR_HOURS_QUALITY_EXCLUSION_REASON = "维修工时超过99%分位数，疑似数据异常"
REPAIR_HOURS_EXCEEDS_ATTENDANCE_REASON = "repairHoursExceedsAttendanceHours"
REPAIR_HOURS_EXCEEDS_ATTENDANCE_EXCLUSION_REASON = "维修工时超过出勤工时，疑似数据异常"

NEW_HOUR_MONTHS = {index + 1: month for index, month in enumerate(MONTH_NAMES)}


def month_order_index(month):
    try:
        return MONTH_ORDER.index(month)
    except ValueError:
        return -1

YEAR_SOURCES = {
    2025: {
        "attendance": "tools_101_2025_attendance_hours.xlsx",
        "order": "tools_101_2025_order_count.xlsx",
        "repair": "tools_101_2025_repair_hours.xlsx",
    },
    2026: {
        "legacyAttendance": "tools_101_2026_legacy_attendance_hours.xlsx",
        "attendance": "tools_101_2026_attendance_hours.xlsx",
        "order": "tools_101_2026_order_count.xlsx",
        "repair": "tools_101_2026_repair_hours.xlsx",
    },
}

AC_SOURCES = {
    2025: {
        "attendance": "ac_103-104_2025_attendance_hours.xlsx",
        "order": "ac_103-104_2025_order_count.xlsx",
        "repair": "ac_103-104_2025_repair_hours.xlsx",
    },
    2026: {
        "attendance": "ac_103-104_2026_attendance_hours.xlsx",
        "order": "ac_103-104_2026_order_count.xlsx",
        "repair": "ac_103-104_2026_repair_hours.xlsx",
    },
}

RELIABILITY_SOURCES = {
    2025: {
        "reactionTimeHours": "tools_101_2025_reaction_time_detail.xlsx",
    },
    2026: {
        "reactionTimeHours": "tools_101_2026_reaction_time_detail.xlsx",
    },
}

TEF31_32_PM_DATA_DIR = "2022-2025 pi 数据"
TEF_ATTENDANCE_DATA_DIR = "tef313133出勤时间"
TEF3132_ATTENDANCE_DIR = "tef3132"
TEF33_ATTENDANCE_DIR = "tef33"
TRANSFER_HOURS_BUNDLE = "转移工时.zip"
ORDER_COUNT_DATA_DIR = "接单数量"

TEF31_32_ATTENDANCE_DETAIL_FILES = {
    2025: "2025 年TEF31、TEF32员工工时明细 .xlsx",
    2026: "副本2026 年TEF31、TEF32员工工时明细  .xlsx",
}

TEF33_ATTENDANCE_BUNDLE = "tef33_2025-2026_attendance_bundle.zip"
PREFERRED_PM_DUPLICATE_KEYWORDS = {
    (2022, 12): "副本",
}

HOLIDAY_3X_DAYS_BY_MONTH = {
    (2025, 5): {1, 2, 3},
    (2026, 5): {1, 2, 3},
}

NAME_ALIASES = {
    "张晓锋": "张晓峰",
    "谢春锋": "谢春峰",
    "吴剑峰": "吴剑锋",
}

TRANSFER_NAME_ALIASES = {
    "caoguofeng": "曹国锋",
    "caiyouguo": "蔡有国",
    "dongjianbing": "董建兵",
    "dongjianbin": "董建兵",
    "duyebo": "杜业波",
    "fangjian": "方建",
    "gaokaihong": "高凯鸿",
    "hebiao": "何彪",
    "jinfeng": "金峰",
    "kongzhong": "孔忠",
    "laisongjiang": "来宋江",
    "liguohua": "李国华",
    "liutao": "刘涛",
    "liuweibing": "刘卫兵",
    "liuweibin": "刘卫兵",
    "qiuguoliang": "邱国亮",
    "xieshujin": "谢书金",
    "xuhaihua": "许海华",
    "xuwei": "徐伟",
    "yaozhihui": "姚智慧",
    "yangzhengguo": "杨振国",
    "yangzhengjiao": "杨正蛟",
    "yangzhenguo": "杨振国",
    "yubin": "俞斌",
    "yuhuamiao": "虞华苗",
    "zhangle": "张乐",
    "zhengping": "郑平",
}

ORDER_NAME_ALIASES = {
    **TRANSFER_NAME_ALIASES,
    "baoru": "白茹",
    "baoqijun": "包奇军",
    "caoxiangyang": "曹向阳",
    "chenjinhai": "陈金海",
    "chenjiande": "陈建德",
    "chenxuanzhi": "陈轩志",
    "chenwei": "陈伟",
    "duanyangyang": "段洋洋",
    "fangjianqing": "方建青",
    "ganchao": "甘超",
    "gaokaihong": "高凯鸿",
    "gaozhitao": "高志涛",
    "geqiping": "葛其平",
    "hebiao": "何彪",
    "hehongbo": "何洪波",
    "herenhong": "何仁红",
    "huqi": "胡骑",
    "huxiaorong": "胡小荣",
    "huzenghui": "胡增辉",
    "jiangjianbing": "蒋建兵",
    "jiangweiqiang": "蒋卫强",
    "kongdeyi": "孔德义",
    "kongguorong": "孔国荣",
    "laiguoming": "来国明",
    "lei guangzheng": "雷光正",
    "leiguangzheng": "雷光正",
    "lifei": "李飞",
    "lumengyun": "鲁孟云",
    "liutao": "刘涛",
    "liuweibing": "刘卫兵",
    "mengshijie": "孟世杰",
    "oubing": "欧冰",
    "qiuhangdong": "裘杭东",
    "sunbingcheng": "孙兵成",
    "sunjiandong": "孙建东",
    "sunquanzhi": "孙全治",
    "wangchenjie": "王陈杰",
    "wangguoyong": "王国勇",
    "wangjianren": "王建人",
    "wangwei": "王伟",
    "wangxuebing": "王学兵",
    "wangyizhi": "王一志",
    "wu jianfeng": "吴剑锋",
    "wuqiang": "吴强",
    "wuxiaoqing": "吴小青",
    "xietong": "谢逵",
    "xiechunfeng": "谢春峰",
    "xielitong": "谢礼同",
    "xue tengfei": "薛腾飞",
    "xuetengfei": "薛腾飞",
    "yangjianbo": "杨建波",
    "yangzhenguo": "杨正国",
    "yu xiang": "余翔",
    "yuxiang": "余翔",
    "yuanxiaobiao": "袁小彪",
    "zhachongnan": "查中南",
    "zhazhongnan": "查中南",
    "zhangbin": "张斌",
    "zhangsong": "张松",
    "zhangzhigui": "张志贵",
    "zhaodonglun": "赵东伦",
    "zhaotong": "赵彤",
    "zhenxiaojun": "郑晓军",
    "zhengguohua": "郑国华",
    "zhangjuntao": "张军涛",
    "zhangzhigang": "张志刚",
    "zhangxiaofeng": "张晓峰",
    "zhengyuzhong": "郑余忠",
    "zhouxiaofei": "周晓飞",
    "zhouyongding": "周永定",
}

TEAM_ORDER_NAME_KEYS = {
    "mfm维修团队",
    "103维修团队",
    "104维修团队",
    "维修团队",
}


def normalize_name(value):
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", "", text).strip()
    return NAME_ALIASES.get(text, text)


def transfer_name_key(value):
    text = "" if value is None else str(value)
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", text).casefold()


def is_valid_name(value):
    name = normalize_name(value)
    return bool(name and name.lower() != "nan" and name not in {"0", "姓名", "Total", "总计", "合计", "小计"})


def clean_number(value):
    if value is None:
        return 0
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0
    if not math.isfinite(number) or pd.isna(number):
        return 0
    return number


def optional_number(value):
    if value is None or pd.isna(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def hour_number(value):
    number = optional_number(value)
    if number is None:
        return 0
    return abs(number)


def round_number(value, digits=4):
    number = clean_number(value)
    return round(number, digits)


def percentile(values, ratio):
    numbers = sorted(float(value) for value in values if optional_number(value) is not None)
    if not numbers:
        return None
    if len(numbers) == 1:
        return numbers[0]
    position = (len(numbers) - 1) * ratio
    lower_index = math.floor(position)
    upper_index = math.ceil(position)
    if lower_index == upper_index:
        return numbers[lower_index]
    weight = position - lower_index
    return numbers[lower_index] * (1 - weight) + numbers[upper_index] * weight


def clean_employee_no(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        if not math.isfinite(value):
            return ""
        return str(int(value)) if float(value).is_integer() else str(value).strip()
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text


def read_export_workbook(filename):
    return pd.read_excel(DATA_DIR / filename, sheet_name="Export")


def is_employee_row(row):
    name = normalize_name(row.get("姓名"))
    shift = row.get("班次")
    if not name or name.lower() == "nan" or name == "Total":
        return False
    if shift is None or str(shift).strip() in {"", "nan", "Total"}:
        return False
    return True


def read_employee_teams(filename=TOOLS_ROSTER_FILE):
    frame = pd.read_excel(DATA_DIR / filename, header=None)
    employees = {}
    shift_order = []

    for _, row in frame.iterrows():
        shift = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
        raw_name = "" if len(row) < 3 or pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()
        if not shift or not raw_name or shift == "班组" or raw_name == "姓名":
            continue
        canonical_name = normalize_name(raw_name)
        if shift not in shift_order:
            shift_order.append(shift)
        employees[canonical_name] = {
            "name": raw_name,
            "canonicalName": canonical_name,
            "shift": shift,
        }

    return employees, shift_order


def frame_values_by_employee(frame):
    values = {}
    for _, row in frame.iterrows():
        if not is_employee_row(row):
            continue
        name = normalize_name(row["姓名"])
        values[name] = {month: clean_number(row.get(month)) for month in EXPORT_MONTH_COLUMNS if month in frame.columns}
    return values


def frame_shifts_by_employee(frame):
    shifts = {}
    for _, row in frame.iterrows():
        if not is_employee_row(row):
            continue
        name = normalize_name(row.get("姓名"))
        shift = "" if pd.isna(row.get("班次")) else str(row.get("班次")).strip()
        if name and shift and shift.lower() != "nan" and name not in shifts:
            shifts[name] = shift
    return shifts


def frame_values_by_key(frame):
    values = {}
    current_area = ""
    current_shift = ""
    for _, row in frame.iterrows():
        area = "" if pd.isna(row.get("Area")) else str(row.get("Area")).strip()
        shift = "" if pd.isna(row.get("Shift")) else str(row.get("Shift")).strip()
        name = normalize_name(row.get("姓名"))

        if area:
            current_area = area
        if shift:
            current_shift = shift
        if not name or name.lower() == "nan" or name == "Total":
            continue
        if not current_area or not current_shift:
            continue

        key = f"{current_area}::{current_shift}::{name}"
        values[key] = {
            month: clean_number(row.get(month))
            for month in MONTH_NAMES
            if month in frame.columns
        }
    return values


def read_new_hours(filename="员工工时.xlsx"):
    frame = pd.read_excel(DATA_DIR / filename, sheet_name="Sheet1")
    frame = frame[pd.to_numeric(frame["SAP"], errors="coerce").notna()].copy()
    values = {}
    for _, row in frame.iterrows():
        name = normalize_name(row["姓名"])
        year = int(clean_number(row.get("年份")) or 0)
        if not year:
            continue
        values[name] = {}
        for column, month in NEW_HOUR_MONTHS.items():
            if column in frame.columns:
                values[name][f"{year} {month}"] = clean_number(row.get(column))
    return values


def values_for_year(frame, year):
    raw_values = frame_values_by_employee(frame)
    return {
        name: {f"{year} {month}": clean_number(month_values.get(month)) for month in MONTH_NAMES}
        for name, month_values in raw_values.items()
    }


def values_for_ac_year(frame, year):
    raw_values = frame_values_by_key(frame)
    return {
        key: {f"{year} {month}": clean_number(month_values.get(month)) for month in MONTH_NAMES}
        for key, month_values in raw_values.items()
    }


def merge_month_values(target, source):
    for name, values in source.items():
        target[name] = {**target.get(name, {}), **values}


def normalize_ac_plant(area, shift):
    area_text = str(area or "").strip()
    shift_text = str(shift or "").strip()
    if area_text in {"103", "104"}:
        return area_text
    if shift_text.startswith("103"):
        return "103"
    if shift_text.startswith("104"):
        return "104"
    return area_text or "AC"


def read_reliability_metrics():
    records = []
    for year, sources in RELIABILITY_SOURCES.items():
        for metric, filename in sources.items():
            frame = read_export_workbook(filename)
            current_value_stream = ""
            for _, row in frame.iterrows():
                raw_value_stream = row.get("Value Stream")
                if raw_value_stream is not None and not pd.isna(raw_value_stream) and str(raw_value_stream).strip():
                    current_value_stream = str(raw_value_stream).strip()

                bps_name = row.get("Corresponding BPS name")
                if bps_name is None or pd.isna(bps_name) or not str(bps_name).strip():
                    continue
                bps_name = str(bps_name).strip()

                for month in MONTH_NAMES:
                    value = optional_number(row.get(month))
                    if value is None:
                        continue
                    records.append(
                        {
                            "metric": metric,
                            "year": year,
                            "month": f"{year} {month}",
                            "monthNumber": MONTH_NAMES.index(month) + 1,
                            "valueStream": current_value_stream,
                            "bpsName": bps_name,
                            "value": round_number(value, 4),
                            "target": round_number(row.get("Target"), 4) if optional_number(row.get("Target")) is not None else None,
                            "sourceFile": f"{RAW_DATA_LABEL}/{filename}",
                        }
                    )
    return records


def workbook_month_from_name(filename):
    match = re.search(r"(\d{1,2})月份维修数据", filename)
    if match:
        month_number = int(match.group(1))
        if 1 <= month_number <= 12:
            return month_number
    match = re.search(r"_(\d{2})_repair_detail_workbook", filename)
    if match:
        month_number = int(match.group(1))
        if 1 <= month_number <= 12:
            return month_number
    return None


def year_from_short_filename(filename):
    match = re.search(r"(\d{2})", str(filename or ""))
    if not match:
        return None
    year = 2000 + int(match.group(1))
    return year if 2020 <= year <= 2030 else None


def year_month_from_attendance_filename(filename):
    match = re.search(r"(20\d{2})\D*([01]\d)", str(filename or ""))
    if not match:
        return None, None
    year = int(match.group(1))
    month_number = int(match.group(2))
    if 1 <= month_number <= 12:
        return year, month_number
    return None, None


def fixed_zip_member_name(filename):
    try:
        return filename.encode("cp437").decode("gbk")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return filename


def transfer_scope_from_sheet(filename, sheet_name):
    path = fixed_zip_member_name(filename)
    if "2023/" in path:
        if sheet_name == "Tools":
            return "Tools"
        if sheet_name == "AC":
            return "AC"
    if sheet_name in {"MFM1", "MFM2", "TEF31", "TEF32"}:
        return "Tools"
    if sheet_name in {"MFM3", "TEF33"}:
        return "AC"
    return ""


def transfer_year_from_filename(filename):
    match = re.search(r"(20\d{2})", fixed_zip_member_name(filename))
    if not match:
        return None
    year = int(match.group(1))
    return year if 2020 <= year <= 2030 else None


def parse_transfer_month(value, fallback_year):
    if isinstance(value, (datetime, date)):
        return value.year, value.month

    text = "" if value is None else str(value).strip()
    if not text:
        return None, None

    match = re.search(r"(20\d{2})\D+(\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month_number = int(match.group(2))
        if 1 <= month_number <= 12:
            return year, month_number

    match = re.search(r"^(\d{1,2})/(\d{1,2})/(20\d{2})$", text)
    if match:
        first, second, year = (int(part) for part in match.groups())
        month_number = second if first > 12 else first
        if 1 <= month_number <= 12:
            return year, month_number

    match = re.search(r"^(\d{1,2})[\-.月]", text)
    if match and fallback_year:
        month_number = int(match.group(1))
        if 1 <= month_number <= 12:
            return fallback_year, month_number

    return None, None


def normalize_transfer_responsible(value):
    raw_name = "" if value is None else str(value).strip()
    name_key = transfer_name_key(raw_name)
    if not raw_name or raw_name.lower() == "nan":
        return "", raw_name
    alias_lookup = {
        **{transfer_name_key(key): value for key, value in ORDER_NAME_ALIASES.items()},
        **{transfer_name_key(key): value for key, value in TRANSFER_NAME_ALIASES.items()},
    }
    if name_key in alias_lookup:
        return normalize_name(alias_lookup[name_key]), raw_name
    if "/" in raw_name or len(re.findall(r"[A-Z][a-z]+", raw_name)) > 2:
        return "", raw_name
    if re.fullmatch(r"[\d.]+", raw_name):
        return "", raw_name
    return normalize_name(raw_name), raw_name


def read_transfer_hour_values(existing_record_keys):
    values = {}
    unmatched = []
    supplemental = []
    matched_hours = 0
    supplemental_hours = 0
    unmatched_hours = 0
    total_hours = 0
    parsed_rows = 0
    bundle_path = DATA_DIR / TRANSFER_HOURS_BUNDLE
    if not bundle_path.exists():
        return values, {
            "sourceExists": False,
            "parsedRows": 0,
            "matchedHours": 0,
            "supplementalHours": 0,
            "unmatchedHours": 0,
            "totalHours": 0,
            "supplemental": [],
            "unmatched": [],
        }

    with ZipFile(bundle_path) as archive:
        for item in archive.infolist():
            basename = item.filename.split("/")[-1]
            if item.is_dir() or basename.startswith("~$") or basename.startswith("."):
                continue
            if not basename.lower().endswith((".xlsx", ".xlsm")):
                continue

            source_file = fixed_zip_member_name(item.filename)
            fallback_year = transfer_year_from_filename(source_file)
            workbook = load_workbook(BytesIO(archive.read(item)), read_only=True, data_only=True)
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(values_only=True)
                header_row_index = None
                column_map = {}
                buffered_rows = []
                for index, row in enumerate(rows, start=1):
                    buffered_rows.append(row)
                    if index > 8:
                        break
                    labels = ["" if value is None else str(value).strip() for value in row]
                    if "日期" in labels and ("提出者" in labels or "责任人" in labels) and any("工时" in label for label in labels):
                        header_row_index = index
                        column_map = {label: column_index for column_index, label in enumerate(labels) if label}
                        break

                if header_row_index is None:
                    continue

                date_column = column_map["日期"]
                owner_column = column_map.get("提出者", column_map.get("责任人"))
                responsible_column = column_map.get("责任人")
                hour_column = next(index for label, index in column_map.items() if "工时" in label)
                scope = transfer_scope_from_sheet(source_file, worksheet.title)
                data_rows = buffered_rows[header_row_index:]
                data_rows.extend(rows)

                for row in data_rows:
                    raw_hours = row[hour_column] if hour_column < len(row) else 0
                    hours = round_number(raw_hours, 4)
                    if hours <= 0:
                        continue

                    parsed_rows += 1
                    total_hours += hours
                    raw_owner_value = row[owner_column] if owner_column is not None and owner_column < len(row) else ""
                    name, raw_name = normalize_transfer_responsible(raw_owner_value)
                    raw_responsible_value = (
                        row[responsible_column]
                        if responsible_column is not None and responsible_column < len(row)
                        else ""
                    )
                    year, month_number = parse_transfer_month(row[date_column] if date_column < len(row) else "", fallback_year)
                    month = f"{year} {MONTH_NAMES[month_number - 1]}" if year and month_number else ""
                    reason = ""
                    if not scope:
                        reason = "unknownScope"
                    elif not month:
                        reason = "unparsedMonth"
                    elif month not in MONTH_ORDER:
                        reason = "outsideDashboardRange"
                    elif not name:
                        reason = "unconfirmedRequester"
                    elif (scope, name, month) not in existing_record_keys:
                        reason = "noEmployeeMonthRecord"

                    if reason:
                        entry = {
                            "reason": reason,
                            "sourceFile": f"{RAW_DATA_LABEL}/{TRANSFER_HOURS_BUNDLE}:{source_file}",
                            "sheet": worksheet.title,
                            "businessArea": scope,
                            "ownerColumn": "提出者" if "提出者" in column_map else "责任人",
                            "rawRequester": raw_name,
                            "rawResponsible": "" if raw_responsible_value is None else str(raw_responsible_value).strip(),
                            "sourceResponsible": "" if raw_responsible_value is None else str(raw_responsible_value).strip(),
                            "mappedName": name,
                            "month": month,
                            "hours": hours,
                        }
                        if reason in SUPPLEMENTAL_TRANSFER_REASONS and scope and month in MONTH_ORDER:
                            supplemental.append(entry)
                            supplemental_hours += hours
                        else:
                            unmatched.append(entry)
                            unmatched_hours += hours
                        continue

                    key = (scope, name, month)
                    values[key] = round_number(values.get(key, 0) + hours, 4)
                    matched_hours += hours
            workbook.close()

    return values, {
        "sourceExists": True,
        "parsedRows": parsed_rows,
        "matchedEmployeeMonthKeys": len(values),
        "matchedHours": round_number(matched_hours, 4),
        "supplementalRows": len(supplemental),
        "supplementalHours": round_number(supplemental_hours, 4),
        "totalHours": round_number(total_hours, 4),
        "ownerColumn": "提出者",
        "representedHours": round_number(matched_hours + supplemental_hours, 4),
        "unmatchedRows": len(unmatched),
        "unmatchedHours": round_number(unmatched_hours, 4),
        "supplemental": supplemental,
        "unmatched": unmatched,
    }


def read_tef3132_attendance_values():
    values = {}
    shifts = {}
    attendance_dir = DATA_DIR / TEF_ATTENDANCE_DATA_DIR / TEF3132_ATTENDANCE_DIR
    if not attendance_dir.exists():
        return values, shifts

    for path in sorted(attendance_dir.glob("*.xlsx")):
        year = year_from_short_filename(path.name)
        if not year:
            continue
        frame = pd.read_excel(path, sheet_name="Export")
        current_shift = ""
        for _, row in frame.iterrows():
            raw_shift = "" if pd.isna(row.get("班次")) else str(row.get("班次")).strip()
            if raw_shift:
                current_shift = raw_shift
            name = normalize_name(row.get("姓名"))
            if not is_valid_name(name):
                continue
            if current_shift:
                shifts.setdefault(name, current_shift)
            month_values = {
                f"{year} {month}": clean_number(row.get(month))
                for month in MONTH_NAMES
                if month in frame.columns
            }
            values[name] = {**values.get(name, {}), **month_values}
    return values, shifts


def read_tef33_attendance_values_by_name():
    values = {}
    attendance_dir = DATA_DIR / TEF_ATTENDANCE_DATA_DIR / TEF33_ATTENDANCE_DIR
    if not attendance_dir.exists():
        return values

    for path in sorted(attendance_dir.iterdir()):
        if not path.name.lower().endswith((".xlsx", ".xlsm")) or path.name.startswith("~$"):
            continue
        year, month_number = year_month_from_attendance_filename(path.name)
        if not year or not month_number:
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            workbook.close()
            continue

        labels = ["" if value is None else str(value).strip() for value in header]
        name_column = next((index for index, label in enumerate(labels) if "Name" in label or "姓名" in label), None)
        attendance_column = next((index for index, label in enumerate(labels) if "Total Working Hours" in label or "总工作小时数" in label), None)
        if name_column is None or attendance_column is None:
            workbook.close()
            continue

        month = f"{year} {MONTH_NAMES[month_number - 1]}"
        for row in rows:
            raw_name = row[name_column] if name_column < len(row) else ""
            name = normalize_name(str(raw_name).split("/")[0])
            if not is_valid_name(name):
                continue
            values.setdefault(name, {})[month] = round_number(row[attendance_column] if attendance_column < len(row) else 0, 4)
        workbook.close()
    return values


def merge_ac_attendance_by_name(ac_attendance_values, ac_keys, attendance_by_name, allowed_months_by_key=None):
    allowed_months_by_key = allowed_months_by_key or {}
    keys_by_name = {}
    for key in ac_keys:
        name = normalize_name(key.split("::", 2)[2])
        keys_by_name.setdefault(name, []).append(key)

    mapped_names = 0
    for name, month_values in attendance_by_name.items():
        keys = keys_by_name.get(name, [])
        if len(keys) != 1:
            continue
        key = keys[0]
        allowed_months = allowed_months_by_key.get(key)
        if allowed_months is not None:
            month_values = {month: value for month, value in month_values.items() if month in allowed_months}
        if not month_values:
            continue
        ac_attendance_values[key] = {**ac_attendance_values.get(key, {}), **month_values}
        mapped_names += 1
    return mapped_names


def pm_data_year_from_name(filename):
    match = re.search(r"(20\d{2})", str(filename or ""))
    return int(match.group(1)) if match else None


def read_pm_hours_from_workbook(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return {}

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = rows[0]
    day_columns = [
        index
        for index, value in enumerate(header)
        if isinstance(value, (int, float)) and 1 <= int(value) <= 31
    ]
    values = {}

    for row in rows[1:]:
        name = normalize_name(row[2] if len(row) > 2 else "")
        if not is_valid_name(name):
            continue
        minutes = sum(clean_number(row[index] if index < len(row) else 0) for index in day_columns)
        values[name] = values.get(name, 0) + minutes / 60

    return values


def merge_pm_workbook_values(target, year, month_number, workbook):
    month = f"{year} {MONTH_NAMES[month_number - 1]}"
    pm01_values = read_pm_hours_from_workbook(workbook, "个人维修时间")
    pm03_values = read_pm_hours_from_workbook(workbook, "个人PM03时间")

    for name in set(pm01_values) | set(pm03_values):
        target.setdefault(name, {}).setdefault(month, {})
        target[name][month]["pm01Hours"] = round_number(pm01_values.get(name, 0), 4)
        target[name][month]["pm03Hours"] = round_number(pm03_values.get(name, 0), 4)


def read_pm_work_detail_values():
    values = {}
    candidates = {}
    pm_data_dir = DATA_DIR / TEF31_32_PM_DATA_DIR
    if not pm_data_dir.exists():
        return values

    for zip_path in sorted(pm_data_dir.glob("*.zip")):
        year = pm_data_year_from_name(zip_path.name)
        if not year:
            continue
        with ZipFile(zip_path) as archive:
            for item in archive.infolist():
                basename = item.filename.split("/")[-1]
                if item.is_dir() or basename.startswith("~$") or basename.startswith("."):
                    continue
                if not basename.lower().endswith((".xlsx", ".xlsm")):
                    continue
                month_number = workbook_month_from_name(basename)
                if not month_number:
                    continue

                key = (year, month_number)
                candidate = (zip_path, item)
                existing = candidates.get(key)
                preferred_keyword = PREFERRED_PM_DUPLICATE_KEYWORDS.get(key)
                if existing is None:
                    candidates[key] = candidate
                elif preferred_keyword:
                    existing_is_preferred = preferred_keyword in existing[1].filename
                    current_is_preferred = preferred_keyword in item.filename
                    if current_is_preferred and not existing_is_preferred:
                        candidates[key] = candidate
                elif "副本" in existing[1].filename and "副本" not in item.filename:
                    candidates[key] = candidate

    for (year, month_number), (zip_path, item) in sorted(candidates.items()):
        with ZipFile(zip_path) as archive:
            workbook = load_workbook(BytesIO(archive.read(item)), read_only=True, data_only=True)
            merge_pm_workbook_values(values, year, month_number, workbook)
            workbook.close()

    return values


def pm_source_file_paths():
    pm_data_dir = DATA_DIR / TEF31_32_PM_DATA_DIR
    if not pm_data_dir.exists():
        return []
    return [
        f"{RAW_DATA_LABEL}/{TEF31_32_PM_DATA_DIR}/{path.name}"
        for path in sorted(pm_data_dir.glob("*.zip"))
        if not path.name.startswith(".")
    ]


def source_file_paths(*parts):
    source_dir = DATA_DIR.joinpath(*parts)
    if not source_dir.exists():
        return []
    return [
        f"{RAW_DATA_LABEL}/{'/'.join(parts)}/{path.name}"
        for path in sorted(source_dir.iterdir())
        if path.is_file() and not path.name.startswith(".") and not path.name.startswith("~$")
    ]


def order_month_from_text(value):
    text = fixed_zip_member_name(str(value or ""))
    match = re.search(r"(?<!\d)0?([1-9]|1[0-2])\s*月", text)
    if not match:
        return None
    month_number = int(match.group(1))
    return month_number if 1 <= month_number <= 12 else None


def order_year_from_path(path):
    match = re.search(r"(20\d{2})", str(path))
    if not match:
        return None
    year = int(match.group(1))
    return year if 2020 <= year <= 2030 else None


def order_plant_from_path(path):
    text = fixed_zip_member_name(str(path))
    for plant in ["101", "103", "104"]:
        if re.search(rf"(^|[^\d]){plant}([^\d]|$)", text):
            return plant
    if re.search(r"\bB?104\b", text, flags=re.IGNORECASE):
        return "104"
    return None


def order_source_label(candidate):
    if candidate.get("member"):
        return f"{RAW_DATA_LABEL}/{candidate['path'].relative_to(DATA_DIR)}:{fixed_zip_member_name(candidate['member'])}"
    return f"{RAW_DATA_LABEL}/{candidate['path'].relative_to(DATA_DIR)}"


def order_workbook_candidates():
    order_dir = DATA_DIR / ORDER_COUNT_DATA_DIR
    if not order_dir.exists():
        return {}

    candidates = {}
    for path in sorted(order_dir.rglob("*")):
        if not path.is_file() or path.name.startswith(".") or path.name.startswith("~$"):
            continue
        suffix = path.suffix.lower()
        if suffix not in {".zip", ".xlsx", ".xlsm"}:
            continue

        if suffix == ".zip":
            with ZipFile(path) as archive:
                for item in archive.infolist():
                    basename = fixed_zip_member_name(item.filename).split("/")[-1]
                    if item.is_dir() or basename.startswith(".") or basename.startswith("~$"):
                        continue
                    if not basename.lower().endswith((".xlsx", ".xlsm")):
                        continue

                    year = order_year_from_path(path)
                    month_number = order_month_from_text(basename)
                    plant = order_plant_from_path(path)
                    if not year or not month_number or not plant:
                        continue
                    key = (plant, year, month_number)
                    candidate = {
                        "path": path,
                        "member": item.filename,
                        "rank": 1,
                        "plant": plant,
                        "year": year,
                        "monthNumber": month_number,
                    }
                    existing = candidates.get(key)
                    if existing is None or candidate["rank"] < existing["rank"]:
                        candidates[key] = candidate
            continue

        year = order_year_from_path(path)
        month_number = order_month_from_text(path.name)
        plant = order_plant_from_path(path)
        if not year or not month_number or not plant:
            continue
        key = (plant, year, month_number)
        candidate = {
            "path": path,
            "member": None,
            "rank": 0,
            "plant": plant,
            "year": year,
            "monthNumber": month_number,
        }
        existing = candidates.get(key)
        if existing is None or candidate["rank"] < existing["rank"]:
            candidates[key] = candidate

    return candidates


def read_order_aliases_from_workbook(workbook):
    aliases = {}
    if "技工名单" not in workbook.sheetnames:
        return aliases

    worksheet = workbook["技工名单"]
    for header_row_index, header in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(8, worksheet.max_row or 0), values_only=True),
        start=1,
    ):
        labels = ["" if value is None else str(value).strip() for value in header]
        technician_column = next((index for index, label in enumerate(labels) if label == "技工名字"), None)
        chinese_name_column = next((index for index, label in enumerate(labels) if label == "姓名"), None)
        if technician_column is None or chinese_name_column is None:
            continue

        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            raw_name = row[technician_column] if technician_column < len(row) else ""
            chinese_name = row[chinese_name_column] if chinese_name_column < len(row) else ""
            key = transfer_name_key(raw_name)
            canonical_name = normalize_name(chinese_name)
            if key and is_valid_name(canonical_name):
                aliases[key] = canonical_name
        break
    return aliases


def read_tef33_order_name_aliases():
    aliases = {}
    attendance_dir = DATA_DIR / TEF_ATTENDANCE_DATA_DIR / TEF33_ATTENDANCE_DIR
    if not attendance_dir.exists():
        return aliases

    for path in sorted(attendance_dir.iterdir()):
        if not path.name.lower().endswith((".xlsx", ".xlsm")) or path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            workbook.close()
            continue

        labels = ["" if value is None else str(value).strip() for value in header]
        name_column = next((index for index, label in enumerate(labels) if "Name" in label or "姓名" in label), None)
        if name_column is None:
            workbook.close()
            continue

        for row in rows:
            raw_name = row[name_column] if name_column < len(row) else ""
            parts = [part.strip() for part in str(raw_name or "").split("/") if part and part.strip()]
            if len(parts) < 2:
                continue
            chinese_name = normalize_name(parts[0])
            english_name = parts[1]
            if is_valid_name(chinese_name):
                aliases[transfer_name_key(english_name)] = chinese_name
        workbook.close()
    return aliases


def normalize_order_responsible(value, local_aliases=None, global_aliases=None):
    raw_name = "" if value is None else str(value).strip()
    name_key = transfer_name_key(raw_name)
    invalid_keys = {"", "nan", "none", "blank", "空白", "na", "n/a", "总计", "grandtotal", "rowlabels", "行标签"}
    if not raw_name or raw_name.lower() == "nan" or name_key in invalid_keys:
        return "", raw_name, "invalid"
    if name_key in TEAM_ORDER_NAME_KEYS:
        return "", raw_name, "teamRow"

    local_aliases = local_aliases or {}
    global_aliases = global_aliases or {}
    alias_lookup = {
        **{transfer_name_key(key): value for key, value in ORDER_NAME_ALIASES.items()},
        **global_aliases,
        **local_aliases,
    }
    if name_key in alias_lookup:
        return normalize_name(alias_lookup[name_key]), raw_name, ""
    if re.search(r"[\u4e00-\u9fff]", raw_name):
        return normalize_name(raw_name), raw_name, ""
    return raw_name, raw_name, "unmappedName"


def is_count_pivot_sheet(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=min(6, worksheet.max_row or 0), values_only=True):
        first_value = row[0] if row else None
        text = "" if first_value is None else str(first_value)
        if "Count of" in text or "计数" in text:
            return True
    return False


def count_sheet_sort_key(worksheet):
    title = worksheet.title.casefold()
    if title in {"work times", "work time2"}:
        return 0
    if "work" in title and "time" in title:
        return 1
    return 2


def parse_order_count_pivot_sheet(worksheet, local_aliases, global_aliases, notes):
    header_row_index = None
    total_column = None

    for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(12, worksheet.max_row or 0), values_only=True), start=1):
        labels = ["" if value is None else str(value).strip() for value in row]
        first_label = labels[0] if labels else ""
        total_candidates = [
            column_index
            for column_index, label in enumerate(labels)
            if label in {"总计", "Grand Total", "Total"}
        ]
        if first_label in {"行标签", "Row Labels"} and total_candidates:
            header_row_index = index
            total_column = total_candidates[-1]
            break

    if header_row_index is None or total_column is None:
        return {}

    values = {}
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        raw_name = row[0] if row else ""
        name, original_name, reason = normalize_order_responsible(raw_name, local_aliases, global_aliases)
        if reason in {"invalid", "teamRow"}:
            if reason == "teamRow":
                notes["skippedTeamRows"] = notes.get("skippedTeamRows", 0) + 1
            continue
        count = clean_number(row[total_column] if total_column < len(row) else 0)
        if count <= 0:
            continue
        values[name] = round_number(values.get(name, 0) + count, 4)
        if reason:
            notes.setdefault("unmappedNames", set()).add(original_name)
    return values


def read_order_counts_from_pivot(workbook, local_aliases, global_aliases, notes):
    worksheets = sorted(
        [worksheet for worksheet in workbook.worksheets if is_count_pivot_sheet(worksheet)],
        key=count_sheet_sort_key,
    )
    for worksheet in worksheets:
        values = parse_order_count_pivot_sheet(worksheet, local_aliases, global_aliases, notes)
        if values:
            return values, worksheet.title
    return {}, ""


def find_order_detail_header(worksheet):
    for header_row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(12, worksheet.max_row or 0), values_only=True),
        start=1,
    ):
        labels = ["" if value is None else str(value).strip() for value in row]
        technician_column = next((index for index, label in enumerate(labels) if label == "技工"), None)
        if technician_column is None:
            continue
        repair_time_column = next(
            (
                index
                for index, label in enumerate(labels)
                if "维修时间" in label and "开始" not in label and "结束" not in label
            ),
            None,
        )
        order_column = next((index for index, label in enumerate(labels) if label == "订单"), None)
        if repair_time_column is None and order_column is None:
            continue
        return header_row_index, technician_column, repair_time_column, order_column
    return None


def has_present_cell(row, column_index):
    if column_index is None or column_index >= len(row):
        return False
    value = row[column_index]
    if value is None:
        return False
    return str(value).strip() != ""


def parse_order_detail_sheet(worksheet, local_aliases, global_aliases, notes):
    header = find_order_detail_header(worksheet)
    if not header:
        return {}

    header_row_index, technician_column, repair_time_column, order_column = header
    values = {}
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if repair_time_column is not None and not has_present_cell(row, repair_time_column):
            continue
        if repair_time_column is None and order_column is not None and not has_present_cell(row, order_column):
            continue

        raw_name = row[technician_column] if technician_column < len(row) else ""
        name, original_name, reason = normalize_order_responsible(raw_name, local_aliases, global_aliases)
        if reason in {"invalid", "teamRow"}:
            if reason == "teamRow":
                notes["skippedTeamRows"] = notes.get("skippedTeamRows", 0) + 1
            continue
        values[name] = values.get(name, 0) + 1
        if reason:
            notes.setdefault("unmappedNames", set()).add(original_name)
    return values


def read_order_counts_from_details(workbook, local_aliases, global_aliases, notes):
    worksheets = [workbook["本月汇总"]] if "本月汇总" in workbook.sheetnames else workbook.worksheets
    values = {}
    sheet_count = 0
    for worksheet in worksheets:
        sheet_values = parse_order_detail_sheet(worksheet, local_aliases, global_aliases, notes)
        if not sheet_values:
            continue
        sheet_count += 1
        for name, count in sheet_values.items():
            values[name] = values.get(name, 0) + count
    return values, sheet_count


def read_order_counts_from_workbook(workbook, global_aliases, notes):
    local_aliases = read_order_aliases_from_workbook(workbook)
    values, sheet_name = read_order_counts_from_pivot(workbook, local_aliases, global_aliases, notes)
    if values:
        return values, {"method": "pivot", "sheet": sheet_name}

    values, sheet_count = read_order_counts_from_details(workbook, local_aliases, global_aliases, notes)
    return values, {"method": "detail", "sheetCount": sheet_count}


def read_mttr_target_from_rows(rows):
    for row_index, row in enumerate(rows, start=1):
        labels = ["" if value is None else str(value).strip() for value in row]
        if not any(re.search(r"\bMTTR\s*target\b", label, flags=re.IGNORECASE) for label in labels):
            continue
        for value in row:
            number = optional_number(value)
            if number is not None and number > 0:
                return row_index, round_number(number, 4)
        return row_index, None
    return None, None


def mttr_section_start_row(rows):
    for row_index, row in enumerate(rows, start=1):
        labels = ["" if value is None else str(value).strip() for value in row]
        if labels and labels[0].casefold() == "mttr":
            return row_index
    return None


def mttr_total_column(rows, mttr_section_index):
    scan_rows = rows[: max(0, mttr_section_index - 1)]
    for row in reversed(scan_rows):
        for column_index, value in enumerate(row):
            if column_index < 2:
                continue
            label = "" if value is None else str(value).strip().casefold()
            if label in {"total", "grand total", "总计", "合计"}:
                return column_index
    return None


def positive_mttr_number(value):
    number = optional_number(value)
    if number is None or number <= 0:
        return None
    return number


def employee_mttr_from_row(row, total_column):
    end_column = total_column if total_column is not None else len(row)
    day_values = [
        {"columnIndex": index, "value": number}
        for index, number in (
            (index, positive_mttr_number(value))
            for index, value in enumerate(row[2:end_column], start=2)
        )
        if number is not None
    ]
    max_day_value = max((item["value"] for item in day_values), default=None)
    max_day_column = next((item["columnIndex"] for item in day_values if item["value"] == max_day_value), None)

    if total_column is not None and total_column < len(row):
        total_value = positive_mttr_number(row[total_column])
        if total_value is not None:
            return {
                "value": total_value,
                "aggregation": "total",
                "maxDayValue": max_day_value,
                "maxDayColumn": max_day_column,
            }

    if not day_values:
        return {
            "value": None,
            "aggregation": "",
            "maxDayValue": None,
            "maxDayColumn": None,
        }
    return {
        "value": sum(item["value"] for item in day_values) / len(day_values),
        "aggregation": "dailyAverage",
        "maxDayValue": max_day_value,
        "maxDayColumn": max_day_column,
    }


def mttr_quality_issue(mttr_item):
    value = mttr_item.get("value")
    max_day_value = mttr_item.get("maxDayValue")
    if value is not None and value > MTTR_MONTHLY_ERROR_THRESHOLD_MINUTES:
        source_error_minutes = max_day_value if max_day_value is not None and max_day_value > MTTR_DAILY_ERROR_THRESHOLD_MINUTES else value
        return {
            "reason": "sourceMonthlyMttrValueExceeds1440Minutes",
            "sourceErrorMinutes": round_number(source_error_minutes, 4),
            "sourceErrorColumnIndex": mttr_item.get("maxDayColumn") if source_error_minutes == max_day_value else None,
        }
    return None


def read_employee_mttr_from_workbook(workbook, local_aliases, global_aliases, notes):
    if "MTTR" not in workbook.sheetnames:
        return {}

    worksheet = workbook["MTTR"]
    rows = list(worksheet.iter_rows(values_only=True))
    mttr_section_index = mttr_section_start_row(rows)
    target_row_index, target_value = read_mttr_target_from_rows(rows)
    if mttr_section_index is None or target_row_index is None:
        return {}

    total_column = mttr_total_column(rows, mttr_section_index)
    values = {}
    duplicate_values = {}
    for row in rows[mttr_section_index - 1 : target_row_index - 1]:
        raw_name = row[1] if len(row) > 1 else ""
        name, original_name, reason = normalize_order_responsible(raw_name, local_aliases, global_aliases)
        if reason in {"invalid", "teamRow"}:
            if reason == "teamRow":
                notes["skippedTeamRows"] = notes.get("skippedTeamRows", 0) + 1
            continue
        if not is_valid_name(name):
            continue

        mttr_item = employee_mttr_from_row(row, total_column)
        value = mttr_item.get("value")
        aggregation = mttr_item.get("aggregation", "")
        if value is None:
            notes["emptyEmployeeRows"] = notes.get("emptyEmployeeRows", 0) + 1
            continue
        if aggregation == "dailyAverage":
            notes["dailyAverageRows"] = notes.get("dailyAverageRows", 0) + 1
        elif aggregation == "total":
            notes["totalColumnRows"] = notes.get("totalColumnRows", 0) + 1

        if reason:
            notes.setdefault("unmappedNames", set()).add(original_name)
        quality_issue = mttr_quality_issue(mttr_item)
        if quality_issue:
            notes["qualityIssueRows"] = notes.get("qualityIssueRows", 0) + 1
        entry = {
            "value": None if quality_issue else round_number(value, 4),
            "target": round_number(target_value, 4) if target_value is not None else None,
            "aggregation": aggregation,
            "sourceValue": round_number(value, 4),
            "sourceMaxDayValue": round_number(mttr_item["maxDayValue"], 4) if mttr_item.get("maxDayValue") is not None else None,
            "sourceMaxDayColumnIndex": mttr_item.get("maxDayColumn"),
            **({"qualityIssue": quality_issue} if quality_issue else {}),
        }
        if name in values:
            duplicate_values.setdefault(name, [values[name].get("value")]).append(entry.get("value"))
            notes.setdefault("duplicateEmployeeRows", set()).add(name)
        values[name] = entry

    for name, duplicates in duplicate_values.items():
        duplicate_numbers = [value for value in duplicates if value is not None]
        values[name]["value"] = round_number(sum(duplicate_numbers) / len(duplicate_numbers), 4) if duplicate_numbers else None
        values[name]["aggregation"] = "duplicateAverage"
    return values


def read_raw_order_mttr_values(known_ac_keys):
    candidates = order_workbook_candidates()
    global_aliases = read_tef33_order_name_aliases()
    tools_values = {}
    ac_values = {}
    notes = {
        "sourceExists": bool((DATA_DIR / ORDER_COUNT_DATA_DIR).exists()),
        "candidateWorkbooks": len(candidates),
        "workbooks": 0,
        "matchedMonths": 0,
        "matchedEmployeeMonths": 0,
        "missingMttrMonths": [],
        "sourceFiles": [],
        "sourceByPlantMonth": {},
        "skippedTeamRows": 0,
        "emptyEmployeeRows": 0,
        "totalColumnRows": 0,
        "dailyAverageRows": 0,
        "unmappedNames": set(),
        "duplicateEmployeeRows": set(),
    }

    for (plant, year, month_number), candidate in sorted(candidates.items()):
        month = f"{year} {MONTH_NAMES[month_number - 1]}"
        if month not in MONTH_ORDER:
            continue

        if candidate.get("member"):
            with ZipFile(candidate["path"]) as archive:
                workbook = load_workbook(BytesIO(archive.read(candidate["member"])), read_only=True, data_only=True)
                local_aliases = read_order_aliases_from_workbook(workbook)
                mttr_values = read_employee_mttr_from_workbook(workbook, local_aliases, global_aliases, notes)
                workbook.close()
        else:
            workbook = load_workbook(candidate["path"], read_only=True, data_only=True)
            local_aliases = read_order_aliases_from_workbook(workbook)
            mttr_values = read_employee_mttr_from_workbook(workbook, local_aliases, global_aliases, notes)
            workbook.close()

        notes["workbooks"] += 1
        source_label = order_source_label(candidate)
        notes["sourceFiles"].append(source_label)
        plant_month_key = f"{plant}::{month}"
        if not mttr_values:
            notes["missingMttrMonths"].append(plant_month_key)
            continue

        for name, item in mttr_values.items():
            raw_value = item.get("value")
            value = clean_number(raw_value)
            if raw_value is None and item.get("qualityIssue"):
                value = None
            if (value is not None and value <= 0) or not is_valid_name(name):
                continue
            target = item.get("target")
            payload = {
                "value": round_number(value, 4) if value is not None else None,
                "target": round_number(target, 4) if target is not None else None,
                "sourceFile": source_label,
                "aggregation": item.get("aggregation", ""),
            }
            if item.get("qualityIssue"):
                if item.get("sourceValue") is not None:
                    payload["sourceValue"] = round_number(item["sourceValue"], 4)
                if item.get("sourceMaxDayValue") is not None:
                    payload["sourceMaxDayValue"] = round_number(item["sourceMaxDayValue"], 4)
                if item.get("sourceMaxDayColumnIndex") is not None:
                    payload["sourceMaxDayColumnIndex"] = item["sourceMaxDayColumnIndex"]
                payload["qualityIssue"] = item["qualityIssue"]
            if plant == "101":
                tools_values.setdefault(name, {})[month] = payload
                continue
            key = resolve_ac_order_key(plant, name, known_ac_keys)
            ac_values.setdefault(key, {})[month] = payload
        notes["sourceByPlantMonth"][plant_month_key] = source_label

    notes["matchedMonths"] = len(notes["sourceByPlantMonth"])
    notes["toolsEmployeeNames"] = len(tools_values)
    notes["acEmployeeKeys"] = len(ac_values)
    notes["toolsMonths"] = sum(len(values) for values in tools_values.values())
    notes["acMonths"] = sum(len(values) for values in ac_values.values())
    notes["matchedEmployeeMonths"] = notes["toolsMonths"] + notes["acMonths"]
    notes["sourceFiles"] = sorted(notes["sourceFiles"])
    notes["unmappedNames"] = sorted(notes["unmappedNames"])
    notes["duplicateEmployeeRows"] = sorted(notes["duplicateEmployeeRows"])
    return tools_values, ac_values, notes


def reliability_for_employee_month(mttr_values, employee_key, month):
    mttr = mttr_values.get(employee_key, {}).get(month)
    if not mttr:
        return None
    result = {
        "mttrMinutes": mttr.get("value"),
        "mttrTargetMinutes": mttr.get("target"),
        "mttrDataSource": mttr.get("sourceFile", ""),
    }
    if mttr.get("sourceValue") is not None:
        result["mttrSourceValue"] = mttr["sourceValue"]
    if mttr.get("sourceMaxDayValue") is not None:
        result["mttrSourceMaxDayValue"] = mttr["sourceMaxDayValue"]
    if mttr.get("sourceMaxDayColumnIndex") is not None:
        result["mttrSourceMaxDayColumnIndex"] = mttr["sourceMaxDayColumnIndex"]
    if mttr.get("qualityIssue"):
        result["mttrQualityIssue"] = mttr["qualityIssue"]
    return result


def merge_repair_hours_quality_issue(existing_issue, incoming_issue):
    if not existing_issue:
        return incoming_issue
    if not incoming_issue:
        return existing_issue
    reason_order = [existing_issue.get("reason"), incoming_issue.get("reason")]
    return {
        **existing_issue,
        **incoming_issue,
        "reason": ";".join(reason for reason in reason_order if reason),
        "sourceValue": max(clean_number(existing_issue.get("sourceValue")), clean_number(incoming_issue.get("sourceValue"))),
        "threshold": existing_issue.get("threshold") or incoming_issue.get("threshold"),
        "percentile": existing_issue.get("percentile") or incoming_issue.get("percentile"),
        "attendanceHours": incoming_issue.get("attendanceHours") or existing_issue.get("attendanceHours"),
    }


def repair_hours_quality_exclusion_reason(issue):
    reason = issue.get("reason", "")
    reasons = []
    if REPAIR_HOURS_QUALITY_REASON in reason:
        reasons.append(REPAIR_HOURS_QUALITY_EXCLUSION_REASON)
    if REPAIR_HOURS_EXCEEDS_ATTENDANCE_REASON in reason:
        reasons.append(REPAIR_HOURS_EXCEEDS_ATTENDANCE_EXCLUSION_REASON)
    return "；".join(reasons) or REPAIR_HOURS_QUALITY_EXCLUSION_REASON


def repair_hours_quality_issues(ac_repair_values, ac_attendance_values):
    samples = [
        clean_number(value)
        for month_values in ac_repair_values.values()
        for value in month_values.values()
        if clean_number(value) > 0
    ]
    threshold = percentile(samples, REPAIR_HOURS_QUALITY_PERCENTILE)
    issues = {}
    issue_records = []
    if threshold is not None:
        for key, month_values in ac_repair_values.items():
            for month, value in month_values.items():
                repair_hours = clean_number(value)
                if repair_hours <= threshold:
                    continue
                issue = {
                    "reason": REPAIR_HOURS_QUALITY_REASON,
                    "sourceValue": round_number(repair_hours, 4),
                    "threshold": round_number(threshold, 4),
                    "percentile": REPAIR_HOURS_QUALITY_PERCENTILE,
                }
                issues.setdefault(key, {})[month] = merge_repair_hours_quality_issue(issues.get(key, {}).get(month), issue)
                issue_records.append({
                    "reason": REPAIR_HOURS_QUALITY_REASON,
                    "employeeKey": f"AC::{key}",
                    "employeeName": key.split("::", 2)[2],
                    "month": month,
                    "repairHours": round_number(repair_hours, 4),
                    "threshold": round_number(threshold, 4),
                })

    attendance_issue_records = []
    for key, month_values in ac_repair_values.items():
        for month, value in month_values.items():
            repair_hours = clean_number(value)
            attendance_hours = clean_number(ac_attendance_values.get(key, {}).get(month, 0))
            if attendance_hours <= 0 or repair_hours <= attendance_hours:
                continue
            issue = {
                "reason": REPAIR_HOURS_EXCEEDS_ATTENDANCE_REASON,
                "sourceValue": round_number(repair_hours, 4),
                "threshold": round_number(attendance_hours, 4),
                "attendanceHours": round_number(attendance_hours, 4),
            }
            issues.setdefault(key, {})[month] = merge_repair_hours_quality_issue(issues.get(key, {}).get(month), issue)
            attendance_issue_records.append({
                "reason": REPAIR_HOURS_EXCEEDS_ATTENDANCE_REASON,
                "employeeKey": f"AC::{key}",
                "employeeName": key.split("::", 2)[2],
                "month": month,
                "attendanceHours": round_number(attendance_hours, 4),
                "repairHours": round_number(repair_hours, 4),
                "threshold": round_number(attendance_hours, 4),
            })

    all_issue_records = []
    for key, month_values in issues.items():
        for month, issue in month_values.items():
            all_issue_records.append({
                "reason": issue.get("reason", ""),
                "employeeKey": f"AC::{key}",
                "employeeName": key.split("::", 2)[2],
                "month": month,
                "attendanceHours": issue.get("attendanceHours"),
                "repairHours": round_number(issue.get("sourceValue"), 4),
                "threshold": issue.get("threshold"),
                "percentile": issue.get("percentile"),
            })
    all_issue_records = sorted(
        all_issue_records,
        key=lambda item: (-item["repairHours"], item["month"], item["employeeName"], item["reason"]),
    )

    return issues, {
        "percentile": REPAIR_HOURS_QUALITY_PERCENTILE,
        "sampleCount": len(samples),
        "threshold": round_number(threshold, 4) if threshold is not None else None,
        "p99IssueCount": len(issue_records),
        "attendanceIssueCount": len(attendance_issue_records),
        "issueCount": len(all_issue_records),
        "issueRecords": all_issue_records,
    }


def resolve_ac_order_key(plant, employee_name, known_ac_keys):
    matching_keys = [
        key
        for key in known_ac_keys
        if normalize_ac_plant(*key.split("::", 2)[:2]) == str(plant)
        and normalize_name(key.split("::", 2)[2]) == employee_name
    ]
    if len(matching_keys) == 1:
        return matching_keys[0]
    return f"{plant}::历史::{employee_name}"


def merge_ac_order_values(target, source):
    for key, values in source.items():
        target[key] = {**target.get(key, {}), **values}


def read_raw_order_count_values(known_ac_keys):
    tools_values = {}
    ac_values = {}
    candidates = order_workbook_candidates()
    global_aliases = read_tef33_order_name_aliases()
    notes = {
        "sourceExists": bool((DATA_DIR / ORDER_COUNT_DATA_DIR).exists()),
        "workbooks": 0,
        "pivotWorkbooks": 0,
        "detailFallbackWorkbooks": 0,
        "skippedTeamRows": 0,
        "unmappedNames": set(),
        "sourceFiles": [],
    }

    for (plant, year, month_number), candidate in sorted(candidates.items()):
        month = f"{year} {MONTH_NAMES[month_number - 1]}"
        if month not in MONTH_ORDER:
            continue

        if candidate.get("member"):
            with ZipFile(candidate["path"]) as archive:
                workbook = load_workbook(BytesIO(archive.read(candidate["member"])), read_only=True, data_only=True)
                counts, method = read_order_counts_from_workbook(workbook, global_aliases, notes)
                workbook.close()
        else:
            workbook = load_workbook(candidate["path"], read_only=True, data_only=True)
            counts, method = read_order_counts_from_workbook(workbook, global_aliases, notes)
            workbook.close()

        notes["workbooks"] += 1
        notes["sourceFiles"].append(order_source_label(candidate))
        if method.get("method") == "pivot":
            notes["pivotWorkbooks"] += 1
        else:
            notes["detailFallbackWorkbooks"] += 1

        for name, count in counts.items():
            if clean_number(count) <= 0 or not is_valid_name(name):
                continue
            if plant == "101":
                tools_values.setdefault(name, {})[month] = round_number(count, 4)
                continue
            key = resolve_ac_order_key(plant, name, known_ac_keys)
            ac_values.setdefault(key, {})[month] = round_number(count, 4)

    notes["candidateWorkbooks"] = len(candidates)
    notes["sourceFiles"] = sorted(notes["sourceFiles"])
    notes["unmappedNames"] = sorted(notes["unmappedNames"])
    notes["toolsEmployeeNames"] = len(tools_values)
    notes["acEmployeeKeys"] = len(ac_values)
    notes["toolsMonths"] = sum(len(values) for values in tools_values.values())
    notes["acMonths"] = sum(len(values) for values in ac_values.values())
    return tools_values, ac_values, notes


def attendance_month_from_sheet_name(sheet_name):
    match = re.search(r"(\d{1,2})\s*月", str(sheet_name or ""))
    if not match:
        return None
    month_number = int(match.group(1))
    if 1 <= month_number <= 12:
        return month_number
    return None


def should_use_attendance_sheet(sheet_name):
    name = str(sheet_name or "")
    if name == "Sheet1":
        return False
    return "考勤明细" not in name


def read_attendance_detail_values():
    values = {}
    for year, filename in TEF31_32_ATTENDANCE_DETAIL_FILES.items():
        workbook = load_workbook(DATA_DIR / filename, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            if not should_use_attendance_sheet(worksheet.title):
                continue
            month_number = attendance_month_from_sheet_name(worksheet.title)
            if not month_number:
                continue

            header = [cell.value for cell in worksheet[1]]
            column_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
            required_columns = {"姓名", "周末加班", "工作日加班", "请假时间"}
            if not required_columns.issubset(column_map):
                continue

            month = f"{year} {MONTH_NAMES[month_number - 1]}"
            holiday_day_columns = [
                column_map[str(day)]
                for day in HOLIDAY_3X_DAYS_BY_MONTH.get((year, month_number), set())
                if str(day) in column_map
            ]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                name = normalize_name(row[column_map["姓名"]] if column_map["姓名"] < len(row) else "")
                if not is_valid_name(name):
                    continue
                weekday_overtime = hour_number(row[column_map["工作日加班"]] if column_map["工作日加班"] < len(row) else 0)
                weekend_overtime = hour_number(row[column_map["周末加班"]] if column_map["周末加班"] < len(row) else 0)
                leave_hours = hour_number(row[column_map["请假时间"]] if column_map["请假时间"] < len(row) else 0)
                holiday_overtime = 0
                if name == "刘斌":
                    holiday_overtime = sum(hour_number(row[index] if index < len(row) else 0) for index in holiday_day_columns)
                values.setdefault(name, {})[month] = {
                    "overtime15Hours": round_number(weekday_overtime, 4),
                    "overtime20Hours": round_number(weekend_overtime, 4),
                    "overtime30Hours": round_number(holiday_overtime, 4),
                    "leaveHours": round_number(leave_hours, 4),
                    "annualLeaveHours": round_number(leave_hours, 4),
                    "sickLeaveHours": 0,
                    "overtimeDetailSource": f"{RAW_DATA_LABEL}/{filename}",
                }
    return values


def read_employee_numbers_from_attendance_details():
    numbers = {}
    for filename in TEF31_32_ATTENDANCE_DETAIL_FILES.values():
        workbook = load_workbook(DATA_DIR / filename, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            if worksheet.title == "Sheet1":
                continue
            header = [cell.value for cell in worksheet[1]]
            column_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
            name_column = column_map.get("姓名")
            employee_no_column = next(
                (
                    index
                    for label, index in column_map.items()
                    if "SAP" in label.upper() and "工号" in label
                ),
                None,
            )
            if name_column is None or employee_no_column is None:
                continue
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                name = normalize_name(row[name_column] if name_column < len(row) else "")
                employee_no = clean_employee_no(row[employee_no_column] if employee_no_column < len(row) else "")
                if is_valid_name(name) and employee_no:
                    numbers.setdefault(name, employee_no)
    return numbers


def read_employee_numbers_from_workbook(workbook, numbers):
    for worksheet in workbook.worksheets:
        max_scan_rows = min(worksheet.max_row or 0, 40)
        for header_row_index, header_row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
            start=1,
        ):
            labels = ["" if value is None else str(value).strip() for value in header_row]
            name_columns = [
                index
                for index, label in enumerate(labels)
                if label in {"姓名", "员工姓名", "Name"} or "姓名" in label
            ]
            employee_no_columns = [
                index
                for index, label in enumerate(labels)
                if label and ("SAP" in label.upper() or "工号" in label or "员工号" in label or "人员编号" in label)
            ]
            if not name_columns or not employee_no_columns:
                continue

            for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                for name_column in name_columns:
                    name = normalize_name(row[name_column] if name_column < len(row) else "")
                    if not is_valid_name(name):
                        continue
                    for employee_no_column in employee_no_columns:
                        employee_no = clean_employee_no(row[employee_no_column] if employee_no_column < len(row) else "")
                        if employee_no:
                            numbers.setdefault(name, employee_no)
                            break


def read_employee_numbers_from_tef33_bundle():
    numbers = {}
    bundle_path = DATA_DIR / TEF33_ATTENDANCE_BUNDLE
    if not bundle_path.exists():
        return numbers

    with ZipFile(bundle_path) as archive:
        for item in archive.infolist():
            if item.is_dir() or item.filename.split("/")[-1].startswith("~$"):
                continue
            if not item.filename.lower().endswith((".xlsx", ".xlsm")):
                continue
            workbook = load_workbook(BytesIO(archive.read(item)), read_only=True, data_only=True)
            read_employee_numbers_from_workbook(workbook, numbers)
            workbook.close()
    return numbers


def make_record(
    employee,
    month,
    attendance_hours,
    order_count,
    repair_hours,
    pm_values=None,
    transfer_hours=0,
    attendance_detail=None,
    is_total=False,
    reliability=None,
):
    shift = employee["shift"]
    employee_name = "Total" if is_total else employee["name"]
    employee_key = f"{shift}::{'Total' if is_total else employee_name}"
    order_efficiency = order_count / attendance_hours if attendance_hours > 0 else 0
    pm_values = pm_values or {}
    has_pm01_detail = "pm01Hours" in pm_values
    pm01_hours = round_number(pm_values.get("pm01Hours") if has_pm01_detail else repair_hours, 4)
    pm03_hours = round_number(pm_values.get("pm03Hours", 0), 4)
    transfer_hours = round_number(transfer_hours, 4)
    repair_efficiency = (pm01_hours + pm03_hours + transfer_hours) / attendance_hours if attendance_hours > 0 else 0
    repair_hours = pm01_hours if has_pm01_detail else repair_hours
    attendance_detail = attendance_detail or {}
    overtime15_hours = round_number(attendance_detail.get("overtime15Hours"), 4)
    overtime20_hours = round_number(attendance_detail.get("overtime20Hours"), 4)
    overtime30_hours = round_number(attendance_detail.get("overtime30Hours"), 4)
    leave_hours = round_number(attendance_detail.get("leaveHours"), 4)
    annual_leave_hours = round_number(attendance_detail.get("annualLeaveHours", leave_hours), 4)
    sick_leave_hours = round_number(attendance_detail.get("sickLeaveHours"), 4)
    overtime_total_hours = round_number(overtime15_hours + overtime20_hours, 4)
    composite_hours = round_number(overtime_total_hours - leave_hours, 4)

    record = {
        "id": f"{employee_key}::{month}",
        "businessArea": employee.get("businessArea", "Tools"),
        "plant": employee.get("plant", "101"),
        "shift": shift,
        "employeeKey": employee_key,
        "employeeName": employee_name,
        "employeeNo": "" if is_total else employee.get("employeeNo", ""),
        "month": month,
        "isTotal": is_total,
        "attendanceHours": round_number(attendance_hours, 1),
        "orderCount": round_number(order_count, 4),
        "repairHours": round_number(repair_hours, 4),
        "repairEfficiency": round_number(repair_efficiency, 4),
        "orderEfficiency": round_number(order_efficiency, 4),
        "computedRepairEfficiency": round_number(repair_efficiency, 4),
        "overtime15Hours": overtime15_hours,
        "overtime20Hours": overtime20_hours,
        "overtime30Hours": overtime30_hours,
        "overtimeTotalHours": overtime_total_hours,
        "holidayOvertimeHours": overtime30_hours,
        "leaveHours": leave_hours,
        "annualLeaveHours": annual_leave_hours,
        "sickLeaveHours": sick_leave_hours,
        "compositeHours": composite_hours,
    }
    if attendance_detail.get("overtimeDetailSource"):
        record["overtimeDetailSource"] = attendance_detail["overtimeDetailSource"]
        record["leaveDataSource"] = attendance_detail["overtimeDetailSource"]
    for field in ["department", "employmentStatus", "isHistoricalEmployee"]:
        if employee.get(field) is not None:
            record[field] = employee[field]
    record["pm01Hours"] = pm01_hours
    record["pm03Hours"] = pm03_hours
    record["transferHours"] = transfer_hours
    if transfer_hours > 0:
        record["transferDataSource"] = f"{RAW_DATA_LABEL}/{TRANSFER_HOURS_BUNDLE}"
    record["pmDataSource"] = "tef31-32_pm_detail_workbooks" if has_pm01_detail else "repair_hours_as_pm01"
    if pm_values:
        record["pm03DataSource"] = "tef31-32_pm_detail_workbooks"
    reliability = reliability or {}
    mttr_minutes = optional_number(reliability.get("mttrMinutes"))
    if mttr_minutes is not None:
        record["mttrMinutes"] = round_number(mttr_minutes, 4)
        if reliability.get("mttrTargetMinutes") is not None:
            record["mttrTargetMinutes"] = round_number(reliability.get("mttrTargetMinutes"), 4)
        if reliability.get("mttrDataSource"):
            record["mttrDataSource"] = reliability["mttrDataSource"]
    elif reliability.get("mttrQualityIssue"):
        if reliability.get("mttrTargetMinutes") is not None:
            record["mttrTargetMinutes"] = round_number(reliability.get("mttrTargetMinutes"), 4)
        if reliability.get("mttrDataSource"):
            record["mttrDataSource"] = reliability["mttrDataSource"]
    if reliability.get("mttrSourceValue") is not None:
        record["mttrSourceValue"] = round_number(reliability["mttrSourceValue"], 4)
    if reliability.get("mttrSourceMaxDayValue") is not None:
        record["mttrSourceMaxDayValue"] = round_number(reliability["mttrSourceMaxDayValue"], 4)
    if reliability.get("mttrSourceMaxDayColumnIndex") is not None:
        record["mttrSourceMaxDayColumnIndex"] = reliability["mttrSourceMaxDayColumnIndex"]
    if reliability.get("mttrQualityIssue"):
        record["mttrQualityIssue"] = reliability["mttrQualityIssue"]
    return record


def make_ac_record(
    key,
    month,
    attendance_hours,
    order_count,
    repair_hours,
    transfer_hours=0,
    employee_numbers=None,
    reliability=None,
    repair_hours_quality_issue=None,
):
    area, shift, employee_name = key.split("::", 2)
    plant = normalize_ac_plant(area, shift)
    employee_key = f"AC::{area}::{shift}::{employee_name}"
    source_repair_hours = round_number(repair_hours, 4)
    repair_hours_quality_issue = repair_hours_quality_issue or {}
    source_repair_efficiency = (source_repair_hours + transfer_hours) / attendance_hours if attendance_hours > 0 else 0
    if repair_hours_quality_issue:
        repair_hours = 0
    pm01_hours = round_number(repair_hours, 4)
    pm03_hours = 0
    transfer_hours = round_number(transfer_hours, 4)
    repair_efficiency = (pm01_hours + pm03_hours + transfer_hours) / attendance_hours if attendance_hours > 0 else 0
    order_efficiency = order_count / attendance_hours if attendance_hours > 0 else 0
    employee_numbers = employee_numbers or {}
    record = {
        "id": f"{employee_key}::{month}",
        "businessArea": "AC",
        "plant": plant,
        "shift": shift,
        "employeeKey": employee_key,
        "employeeName": employee_name,
        "employeeNo": employee_numbers.get(normalize_name(employee_name), ""),
        "month": month,
        "isTotal": False,
        "attendanceHours": round_number(attendance_hours, 1),
        "orderCount": round_number(order_count, 4),
        "repairHours": round_number(repair_hours, 4),
        "repairEfficiency": round_number(repair_efficiency, 4),
        "orderEfficiency": round_number(order_efficiency, 4),
        "computedRepairEfficiency": round_number(repair_efficiency, 4),
        "pm01Hours": pm01_hours,
        "pm03Hours": pm03_hours,
        "transferHours": transfer_hours,
        "pmDataSource": "ac_repair_hours_as_pm01",
        **({"transferDataSource": f"{RAW_DATA_LABEL}/{TRANSFER_HOURS_BUNDLE}"} if transfer_hours > 0 else {}),
    }
    if repair_hours_quality_issue:
        record.update({
            "repairHoursSourceValue": round_number(repair_hours_quality_issue.get("sourceValue", source_repair_hours), 4),
            "pm01HoursSourceValue": round_number(repair_hours_quality_issue.get("sourceValue", source_repair_hours), 4),
            "repairEfficiencySourceValue": round_number(source_repair_efficiency, 4),
            "repairHoursQualityThreshold": round_number(repair_hours_quality_issue.get("threshold"), 4),
            "repairHoursQualityPercentile": repair_hours_quality_issue.get("percentile"),
            "repairHoursQualityReason": repair_hours_quality_issue.get("reason", REPAIR_HOURS_QUALITY_REASON),
            "repairHoursQualityIssue": repair_hours_quality_issue,
            "excludeFromAverages": True,
            "averageExclusionReason": repair_hours_quality_exclusion_reason(repair_hours_quality_issue),
        })
    reliability = reliability or {}
    mttr_minutes = optional_number(reliability.get("mttrMinutes"))
    if mttr_minutes is not None:
        record["mttrMinutes"] = round_number(mttr_minutes, 4)
        if reliability.get("mttrTargetMinutes") is not None:
            record["mttrTargetMinutes"] = round_number(reliability.get("mttrTargetMinutes"), 4)
        if reliability.get("mttrDataSource"):
            record["mttrDataSource"] = reliability["mttrDataSource"]
    elif reliability.get("mttrQualityIssue"):
        if reliability.get("mttrTargetMinutes") is not None:
            record["mttrTargetMinutes"] = round_number(reliability.get("mttrTargetMinutes"), 4)
        if reliability.get("mttrDataSource"):
            record["mttrDataSource"] = reliability["mttrDataSource"]
    if reliability.get("mttrSourceValue") is not None:
        record["mttrSourceValue"] = round_number(reliability["mttrSourceValue"], 4)
    if reliability.get("mttrSourceMaxDayValue") is not None:
        record["mttrSourceMaxDayValue"] = round_number(reliability["mttrSourceMaxDayValue"], 4)
    if reliability.get("mttrSourceMaxDayColumnIndex") is not None:
        record["mttrSourceMaxDayColumnIndex"] = reliability["mttrSourceMaxDayColumnIndex"]
    if reliability.get("mttrQualityIssue"):
        record["mttrQualityIssue"] = reliability["mttrQualityIssue"]
    return record


def supplemental_transfer_display_name(entry):
    mapped_name = normalize_name(entry.get("mappedName"))
    if mapped_name:
        return mapped_name
    raw_name = str(entry.get("rawRequester") or "").strip()
    if raw_name and raw_name.lower() != "nan":
        return raw_name
    return "未确认提出者"


def group_supplemental_transfer_rows(supplemental_rows):
    groups = {}
    for entry in supplemental_rows:
        business_area = entry.get("businessArea")
        month = entry.get("month")
        if business_area not in {"Tools", "AC"} or month not in MONTH_ORDER:
            continue
        display_name = supplemental_transfer_display_name(entry)
        key = (business_area, display_name, month)
        group = groups.setdefault(
            key,
            {
                "businessArea": business_area,
                "displayName": display_name,
                "month": month,
                "hours": 0,
                "sourceRows": [],
                "reasons": set(),
            },
        )
        group["hours"] = round_number(group["hours"] + clean_number(entry.get("hours")), 4)
        group["sourceRows"].append(entry)
        if entry.get("reason"):
            group["reasons"].add(entry["reason"])

    return sorted(
        groups.values(),
        key=lambda item: (month_order_index(item["month"]), item["businessArea"], item["displayName"]),
    )


def supplemental_transfer_metadata(group):
    source_rows = group["sourceRows"]
    source_files = sorted({row.get("sourceFile", "") for row in source_rows if row.get("sourceFile")})
    raw_requesters = sorted(
        {
            str(row.get("rawRequester") or "").strip()
            for row in source_rows
            if str(row.get("rawRequester") or "").strip()
        }
    )
    source_responsibles = sorted(
        {
            str(row.get("rawResponsible") or "").strip()
            for row in source_rows
            if str(row.get("rawResponsible") or "").strip()
        }
    )
    reasons = sorted(group["reasons"])
    return {
        "transferSupplemental": True,
        "transferSupplementalReason": ",".join(reasons),
        "transferSupplementalReasons": reasons,
        "transferSupplementalSourceRows": len(source_rows),
        "transferSupplementalSourceHours": round_number(group["hours"], 4),
        "transferSupplementalRawRequester": raw_requesters,
        "transferSupplementalRawResponsible": source_responsibles,
        "transferSupplementalSourceResponsible": source_responsibles,
        "transferSupplementalSourceFiles": source_files,
    }


def ac_month_activity_score(key, month, ac_attendance_values, ac_order_values, ac_repair_values):
    return (
        clean_number(ac_attendance_values.get(key, {}).get(month, 0))
        + clean_number(ac_order_values.get(key, {}).get(month, 0))
        + clean_number(ac_repair_values.get(key, {}).get(month, 0))
    )


def ac_total_activity_score(key, ac_attendance_values, ac_order_values, ac_repair_values):
    return (
        sum(clean_number(value) for value in ac_attendance_values.get(key, {}).values())
        + sum(clean_number(value) for value in ac_order_values.get(key, {}).values())
        + sum(clean_number(value) for value in ac_repair_values.get(key, {}).values())
    )


def ac_keys_by_employee_name(ac_keys):
    keys_by_name = {}
    for key in ac_keys:
        name = normalize_name(key.split("::", 2)[2])
        keys_by_name.setdefault(name, []).append(key)
    return keys_by_name


def select_ac_transfer_target_keys(transfer_hour_values, ac_keys, ac_attendance_values, ac_order_values, ac_repair_values):
    keys_by_name = ac_keys_by_employee_name(ac_keys)
    target_keys = {}
    for (scope, name, month), hours in transfer_hour_values.items():
        if scope != "AC" or clean_number(hours) <= 0:
            continue
        candidates = keys_by_name.get(name, [])
        active_candidates = [
            key
            for key in candidates
            if ac_month_activity_score(key, month, ac_attendance_values, ac_order_values, ac_repair_values) > 0
        ]
        candidate_pool = active_candidates or candidates
        if not candidate_pool:
            continue
        target_keys[(name, month)] = sorted(
            candidate_pool,
            key=lambda key: (-ac_month_activity_score(key, month, ac_attendance_values, ac_order_values, ac_repair_values), key),
        )[0]
    return target_keys


def select_ac_supplemental_key(name, ac_keys, ac_attendance_values, ac_order_values, ac_repair_values):
    candidates = ac_keys_by_employee_name(ac_keys).get(name, [])
    if candidates:
        return sorted(
            candidates,
            key=lambda key: (-ac_total_activity_score(key, ac_attendance_values, ac_order_values, ac_repair_values), key),
        )[0]
    return f"104::转移补录-AC::{name}"


def build_supplemental_transfer_records(
    supplemental_rows,
    employees,
    tools_employee_shifts,
    employee_numbers,
    ac_keys,
    ac_attendance_values,
    ac_order_values,
    ac_repair_values,
    ac_employee_numbers,
):
    records = []
    for group in group_supplemental_transfer_rows(supplemental_rows):
        transfer_hours = round_number(group["hours"], 4)
        if transfer_hours <= 0:
            continue

        display_name = group["displayName"]
        canonical_name = normalize_name(display_name)
        if group["businessArea"] == "Tools":
            employee = employees.get(canonical_name)
            if employee:
                employee_data = {
                    **employee,
                    "businessArea": "Tools",
                    "plant": "101",
                    "employeeNo": employee_numbers.get(canonical_name, employee.get("employeeNo", "")),
                }
            else:
                employee_data = {
                    "name": display_name,
                    "canonicalName": canonical_name,
                    "shift": tools_employee_shifts.get(canonical_name, "转移补录-Tools"),
                    "businessArea": "Tools",
                    "plant": "101",
                    "employeeNo": employee_numbers.get(canonical_name, ""),
                }
            record = make_record(employee_data, group["month"], 0, 0, 0, transfer_hours=transfer_hours)
        else:
            key = select_ac_supplemental_key(
                canonical_name,
                ac_keys,
                ac_attendance_values,
                ac_order_values,
                ac_repair_values,
            )
            record = make_ac_record(key, group["month"], 0, 0, 0, transfer_hours, ac_employee_numbers)

        record.update(supplemental_transfer_metadata(group))
        records.append(record)
    return records


def has_activity(record):
    return any(
        clean_number(record.get(field)) > 0
        for field in [
            "attendanceHours",
            "orderCount",
            "repairHours",
            "repairEfficiency",
            "orderEfficiency",
            "pm01Hours",
            "pm03Hours",
            "transferHours",
            "overtime15Hours",
            "overtime20Hours",
            "leaveHours",
            "mttrMinutes",
            "mttrSourceValue",
            "repairHoursSourceValue",
        ]
    )


def main():
    source_frames = {}
    attendance_values = {}
    order_values = {}
    repair_values = {}
    new_hour_values = {}
    reliability_metrics = read_reliability_metrics()
    ac_attendance_values = {}
    ac_order_values = {}
    ac_repair_values = {}
    tools_employee_shifts = {}
    tef33_attendance_mapped_names = 0

    for year, sources in YEAR_SOURCES.items():
        attendance_frame = read_export_workbook(sources["legacyAttendance"] if year == 2026 else sources["attendance"])
        order_frame = read_export_workbook(sources["order"])
        repair_frame = read_export_workbook(sources["repair"])
        source_frames[year] = {
            "attendance": attendance_frame,
            "order": order_frame,
            "repair": repair_frame,
        }
        for frame in [attendance_frame, order_frame, repair_frame]:
            for name, shift in frame_shifts_by_employee(frame).items():
                tools_employee_shifts.setdefault(name, shift)
        attendance_values.update({name: {**attendance_values.get(name, {}), **values} for name, values in values_for_year(attendance_frame, year).items()})
        order_values.update({name: {**order_values.get(name, {}), **values} for name, values in values_for_year(order_frame, year).items()})
        repair_values.update({name: {**repair_values.get(name, {}), **values} for name, values in values_for_year(repair_frame, year).items()})

    tef3132_attendance_values, tef3132_shifts = read_tef3132_attendance_values()
    merge_month_values(attendance_values, tef3132_attendance_values)
    for name, shift in tef3132_shifts.items():
        tools_employee_shifts.setdefault(name, shift)

    for year, sources in AC_SOURCES.items():
        attendance_frame = read_export_workbook(sources["attendance"])
        order_frame = read_export_workbook(sources["order"])
        repair_frame = read_export_workbook(sources["repair"])
        ac_attendance_values.update({key: {**ac_attendance_values.get(key, {}), **values} for key, values in values_for_ac_year(attendance_frame, year).items()})
        ac_order_values.update({key: {**ac_order_values.get(key, {}), **values} for key, values in values_for_ac_year(order_frame, year).items()})
        ac_repair_values.update({key: {**ac_repair_values.get(key, {}), **values} for key, values in values_for_ac_year(repair_frame, year).items()})

    ac_key_seed = sorted(set(ac_attendance_values) | set(ac_order_values) | set(ac_repair_values))
    raw_order_values, raw_ac_order_values, raw_order_notes = read_raw_order_count_values(ac_key_seed)
    raw_order_mttr_values, raw_ac_order_mttr_values, raw_order_mttr_notes = read_raw_order_mttr_values(sorted(set(ac_key_seed) | set(raw_ac_order_values)))
    merge_month_values(order_values, raw_order_values)
    merge_ac_order_values(ac_order_values, raw_ac_order_values)

    ac_key_seed = sorted(set(ac_attendance_values) | set(ac_order_values) | set(ac_repair_values) | set(raw_ac_order_mttr_values))
    ac_allowed_months = {
        key: set(ac_attendance_values.get(key, {}))
        | set(ac_order_values.get(key, {}))
        | set(ac_repair_values.get(key, {}))
        | set(raw_ac_order_mttr_values.get(key, {}))
        for key in ac_key_seed
    }
    tef33_attendance_values = read_tef33_attendance_values_by_name()
    tef33_attendance_mapped_names = merge_ac_attendance_by_name(
        ac_attendance_values,
        ac_key_seed,
        tef33_attendance_values,
        ac_allowed_months,
    )

    employees, shift_order = read_employee_teams()
    new_hour_values = read_new_hours(YEAR_SOURCES[2026]["attendance"])
    pm_work_detail_values = read_pm_work_detail_values()
    attendance_detail_values = read_attendance_detail_values()
    employee_numbers = read_employee_numbers_from_attendance_details()
    ac_employee_numbers = read_employee_numbers_from_tef33_bundle()
    for name, employee in employees.items():
        if employee_numbers.get(name):
            employee["employeeNo"] = employee_numbers[name]

    records = []
    metric_names = set(attendance_values) | set(order_values) | set(repair_values) | set(raw_order_mttr_values)
    unmatched_metric_names = sorted(metric_names - set(employees))
    unmatched_new_hour_names = sorted(set(new_hour_values) - set(employees))
    unmatched_pm_work_detail_names = sorted(set(pm_work_detail_values) - set(employees))
    historical_employee_names = sorted((metric_names | set(pm_work_detail_values)) - set(employees))
    historical_employee_records = []
    existing_transfer_record_keys = set()

    for name in set(employees) | set(historical_employee_names):
        for month in MONTH_ORDER:
            attendance_hours = attendance_values.get(name, {}).get(month, 0)
            if new_hour_values.get(name, {}).get(month, 0) > 0:
                attendance_hours = new_hour_values[name][month]
            order_count = order_values.get(name, {}).get(month, 0)
            repair_hours = repair_values.get(name, {}).get(month, 0)
            pm_values = pm_work_detail_values.get(name, {}).get(month) or {}
            attendance_detail = attendance_detail_values.get(name, {}).get(month) or {}
            if any(
                clean_number(value) > 0
                for value in [
                    attendance_hours,
                    order_count,
                    repair_hours,
                    pm_values.get("pm01Hours"),
                    pm_values.get("pm03Hours"),
                    attendance_detail.get("overtime15Hours"),
                    attendance_detail.get("overtime20Hours"),
                    attendance_detail.get("leaveHours"),
                ]
            ):
                existing_transfer_record_keys.add(("Tools", name, month))

    for key in ac_key_seed:
        name = normalize_name(key.split("::", 2)[2])
        for month in MONTH_ORDER:
            if any(
                clean_number(value) > 0
                for value in [
                    ac_attendance_values.get(key, {}).get(month, 0),
                    ac_order_values.get(key, {}).get(month, 0),
                    ac_repair_values.get(key, {}).get(month, 0),
                ]
            ):
                existing_transfer_record_keys.add(("AC", name, month))

    transfer_hour_values, transfer_hour_notes = read_transfer_hour_values(existing_transfer_record_keys)
    ac_repair_hours_quality_issues, ac_repair_hours_quality_notes = repair_hours_quality_issues(
        ac_repair_values,
        ac_attendance_values,
    )

    for month in MONTH_ORDER:
        for shift in shift_order:
            shift_employee_records = []
            for employee in employees.values():
                if employee["shift"] != shift:
                    continue

                name = employee["canonicalName"]
                attendance_hours = attendance_values.get(name, {}).get(month, 0)
                if new_hour_values.get(name, {}).get(month, 0) > 0:
                    attendance_hours = new_hour_values[name][month]

                record = make_record(
                    {**employee, "businessArea": "Tools", "plant": "101"},
                    month,
                    attendance_hours,
                    order_values.get(name, {}).get(month, 0),
                    repair_values.get(name, {}).get(month, 0),
                    pm_work_detail_values.get(name, {}).get(month),
                    transfer_hour_values.get(("Tools", name, month), 0),
                    attendance_detail_values.get(name, {}).get(month),
                    reliability=reliability_for_employee_month(raw_order_mttr_values, name, month),
                )
                if has_activity(record):
                    shift_employee_records.append(record)

            if shift_employee_records:
                records.extend(shift_employee_records)
                total_pm_values = {
                    "pm01Hours": sum(item.get("pm01Hours", 0) for item in shift_employee_records),
                    "pm03Hours": sum(item.get("pm03Hours", 0) for item in shift_employee_records),
                }
                total_transfer_hours = sum(item.get("transferHours", 0) for item in shift_employee_records)
                total_attendance_detail = {
                    "overtime15Hours": sum(item.get("overtime15Hours", 0) for item in shift_employee_records),
                    "overtime20Hours": sum(item.get("overtime20Hours", 0) for item in shift_employee_records),
                    "overtime30Hours": 0,
                    "leaveHours": sum(item.get("leaveHours", 0) for item in shift_employee_records),
                    "annualLeaveHours": sum(item.get("annualLeaveHours", 0) for item in shift_employee_records),
                    "sickLeaveHours": sum(item.get("sickLeaveHours", 0) for item in shift_employee_records),
                    "overtimeDetailSource": "computed: shift total",
                }
                total_employee = {
                    "name": "Total",
                    "canonicalName": "Total",
                    "shift": shift,
                }
                records.append(
                    make_record(
                        {**total_employee, "businessArea": "Tools", "plant": "101"},
                        month,
                        sum(item["attendanceHours"] for item in shift_employee_records),
                        sum(item["orderCount"] for item in shift_employee_records),
                        sum(item["repairHours"] for item in shift_employee_records),
                        total_pm_values,
                        total_transfer_hours,
                        total_attendance_detail,
                        is_total=True,
                    )
                )

    for month in MONTH_ORDER:
        for name in historical_employee_names:
            attendance_hours = attendance_values.get(name, {}).get(month, 0)
            if new_hour_values.get(name, {}).get(month, 0) > 0:
                attendance_hours = new_hour_values[name][month]

            record = make_record(
                {
                    "name": name,
                    "canonicalName": name,
                    "shift": tools_employee_shifts.get(name, "历史"),
                    "businessArea": "Tools",
                    "plant": "101",
                    "employmentStatus": "retired",
                    "isHistoricalEmployee": True,
                    "employeeNo": employee_numbers.get(name, ""),
                },
                month,
                attendance_hours,
                order_values.get(name, {}).get(month, 0),
                repair_values.get(name, {}).get(month, 0),
                pm_work_detail_values.get(name, {}).get(month),
                transfer_hour_values.get(("Tools", name, month), 0),
                attendance_detail_values.get(name, {}).get(month),
                reliability=reliability_for_employee_month(raw_order_mttr_values, name, month),
            )
            if has_activity(record):
                historical_employee_records.append(record)

    records.extend(historical_employee_records)

    ac_keys = sorted(set(ac_attendance_values) | set(ac_order_values) | set(ac_repair_values) | set(raw_ac_order_mttr_values))
    ac_transfer_target_keys = select_ac_transfer_target_keys(
        transfer_hour_values,
        ac_keys,
        ac_attendance_values,
        ac_order_values,
        ac_repair_values,
    )
    for month in MONTH_ORDER:
        for key in ac_keys:
            ac_name = normalize_name(key.split("::", 2)[2])
            transfer_hours = (
                transfer_hour_values.get(("AC", ac_name, month), 0)
                if ac_transfer_target_keys.get((ac_name, month)) == key
                else 0
            )
            record = make_ac_record(
                key,
                month,
                ac_attendance_values.get(key, {}).get(month, 0),
                ac_order_values.get(key, {}).get(month, 0),
                ac_repair_values.get(key, {}).get(month, 0),
                transfer_hours,
                ac_employee_numbers,
                reliability=reliability_for_employee_month(raw_ac_order_mttr_values, key, month),
                repair_hours_quality_issue=ac_repair_hours_quality_issues.get(key, {}).get(month),
            )
            if has_activity(record):
                records.append(record)

    supplemental_transfer_records = build_supplemental_transfer_records(
        transfer_hour_notes.get("supplemental", []),
        employees,
        tools_employee_shifts,
        employee_numbers,
        ac_keys,
        ac_attendance_values,
        ac_order_values,
        ac_repair_values,
        ac_employee_numbers,
    )
    records.extend(supplemental_transfer_records)

    employees_for_filter = [
        {
            "id": f"{employee['shift']}::{employee['name']}",
            "name": employee["name"],
            "businessArea": "Tools",
            "plant": "101",
            "shift": employee["shift"],
            "employeeNo": employee.get("employeeNo", ""),
        }
        for employee in employees.values()
    ]
    historical_employees_for_filter = {
        record["employeeKey"]: {
            "id": record["employeeKey"],
            "name": record["employeeName"],
            "businessArea": record["businessArea"],
            "plant": record["plant"],
            "shift": record["shift"],
            "employeeNo": record.get("employeeNo", ""),
            "employmentStatus": "retired",
        }
        for record in historical_employee_records
    }
    employees_for_filter.extend(historical_employees_for_filter.values())
    supplemental_employees_for_filter = {
        record["employeeKey"]: {
            "id": record["employeeKey"],
            "name": record["employeeName"],
            "businessArea": record["businessArea"],
            "plant": record["plant"],
            "shift": record["shift"],
            "employeeNo": record.get("employeeNo", ""),
            "transferSupplemental": True,
        }
        for record in supplemental_transfer_records
    }
    employees_for_filter.extend(supplemental_employees_for_filter.values())
    employees_for_filter.extend(
        {
            "id": f"AC::{key}",
            "name": key.split("::", 2)[2],
            "businessArea": "AC",
            "plant": normalize_ac_plant(key.split("::", 2)[0], key.split("::", 2)[1]),
            "shift": key.split("::", 2)[1],
            "employeeNo": ac_employee_numbers.get(normalize_name(key.split("::", 2)[2]), ""),
        }
        for key in ac_keys
    )
    deduped_employees_for_filter = []
    seen_employee_filter_ids = set()
    for employee in employees_for_filter:
        if employee["id"] in seen_employee_filter_ids:
            continue
        seen_employee_filter_ids.add(employee["id"])
        deduped_employees_for_filter.append(employee)
    employees_for_filter = deduped_employees_for_filter

    payload = {
        "monthOrder": MONTH_ORDER,
        "records": records,
        "filterOptions": {
            "years": sorted({month.split()[0] for month in MONTH_ORDER}),
            "months": MONTH_ORDER,
            "businessAreas": ["Tools", "AC"],
            "plants": ["101", "103", "104"],
            "shifts": sorted(set(shift_order) | {record["shift"] for record in historical_employee_records + supplemental_transfer_records}),
            "employees": employees_for_filter,
        },
        "sourceFiles": {
            "tools2025AttendanceHours": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2025]['attendance']}",
            "tools2025OrderCount": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2025]['order']}",
            "tools2025RepairHours": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2025]['repair']}",
            "tools2026AttendanceHours": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2026]['attendance']}",
            "tools2026LegacyAttendanceHours": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2026]['legacyAttendance']}",
            "tools2026OrderCount": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2026]['order']}",
            "tools2026RepairHours": f"{RAW_DATA_LABEL}/{YEAR_SOURCES[2026]['repair']}",
            "toolsRoster": f"{RAW_DATA_LABEL}/{TOOLS_ROSTER_FILE}",
            "tools2025ReactionTime": f"{RAW_DATA_LABEL}/{RELIABILITY_SOURCES[2025]['reactionTimeHours']}",
            "tools2026ReactionTime": f"{RAW_DATA_LABEL}/{RELIABILITY_SOURCES[2026]['reactionTimeHours']}",
            "ac2025AttendanceHours": f"{RAW_DATA_LABEL}/{AC_SOURCES[2025]['attendance']}",
            "ac2025OrderCount": f"{RAW_DATA_LABEL}/{AC_SOURCES[2025]['order']}",
            "ac2025RepairHours": f"{RAW_DATA_LABEL}/{AC_SOURCES[2025]['repair']}",
            "ac2026AttendanceHours": f"{RAW_DATA_LABEL}/{AC_SOURCES[2026]['attendance']}",
            "ac2026OrderCount": f"{RAW_DATA_LABEL}/{AC_SOURCES[2026]['order']}",
            "ac2026RepairHours": f"{RAW_DATA_LABEL}/{AC_SOURCES[2026]['repair']}",
            "tef31_32Pm01Pm03Workbooks": pm_source_file_paths(),
            "tef3132AttendanceWorkbooks": source_file_paths(TEF_ATTENDANCE_DATA_DIR, TEF3132_ATTENDANCE_DIR),
            "tef33AttendanceMonthlyFiles": source_file_paths(TEF_ATTENDANCE_DATA_DIR, TEF33_ATTENDANCE_DIR),
            "tef31_32AttendanceDetailFiles": [
                f"{RAW_DATA_LABEL}/{filename}"
                for filename in TEF31_32_ATTENDANCE_DETAIL_FILES.values()
            ],
            "tef33AttendanceBundle": f"{RAW_DATA_LABEL}/{TEF33_ATTENDANCE_BUNDLE}",
            "transferHours": f"{RAW_DATA_LABEL}/{TRANSFER_HOURS_BUNDLE}",
            "rawOrderCounts": raw_order_notes["sourceFiles"],
            "rawOrderMttr": raw_order_mttr_notes["sourceFiles"],
            "repairEfficiency": "computed: (pm01Hours + pm03Hours + transferHours) / attendanceHours",
            "pm01Hours": "computed from sheet 个人维修时间, minutes / 60",
            "pm03Hours": "computed from sheet 个人PM03时间, minutes / 60",
            "overtime15Hours": "from 工作日加班",
            "overtime20Hours": "from 周末加班",
            "overtime30Hours": "holiday overtime display only; not added to total OT because these hours are already included in 周末加班",
            "leaveHours": "from 请假时间",
            "compositeHours": "overtime15Hours + overtime20Hours - leaveHours",
        },
        "reliabilityMetrics": reliability_metrics,
        "generationNotes": {
            "unmatchedMetricNames": unmatched_metric_names,
            "unmatchedNewHourNames": unmatched_new_hour_names,
            "unmatchedPmWorkDetailNames": unmatched_pm_work_detail_names,
            "pmWorkDetailSourceNames": len(pm_work_detail_values),
            "attendanceDetailSourceNames": len(attendance_detail_values),
            "employeeNumberSourceNames": len(employee_numbers),
            "employeeNumberMissingNames": sorted((set(employees) | set(historical_employee_names)) - set(employee_numbers)),
            "historicalEmployeeNames": historical_employee_names,
            "tef3132AttendanceSourceNames": len(tef3132_attendance_values),
            "tef33AttendanceSourceNames": len(tef33_attendance_values),
            "tef33AttendanceMappedNames": tef33_attendance_mapped_names,
            "rawOrderCountSource": raw_order_notes,
            "rawOrderMttrSource": raw_order_mttr_notes,
            "transferHourSource": transfer_hour_notes,
            "acRepairHoursQuality": ac_repair_hours_quality_notes,
            "supplementalTransferRecords": len(supplemental_transfer_records),
            "supplementalTransferRecordHours": round_number(sum(record.get("transferHours", 0) for record in supplemental_transfer_records), 4),
            "acTransferTargetKeys": {
                f"{name}::{month}": key
                for (name, month), key in sorted(ac_transfer_target_keys.items(), key=lambda item: (item[0][0], month_order_index(item[0][1])))
            },
            "nameAliases": NAME_ALIASES,
            "transferNameAliases": TRANSFER_NAME_ALIASES,
            "orderNameAliases": ORDER_NAME_ALIASES,
        },
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.BOSCH_WORKER_MONTHLY_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    employee_records = [item for item in records if not item["isTotal"]]
    print(
        json.dumps(
            {
                "records": len(records),
                "employeeRecords": len(employee_records),
                "months": MONTH_ORDER,
                "shifts": shift_order,
                "employees": len(employees_for_filter),
                "acEmployees": len(ac_keys),
                "reliabilityRecords": len(reliability_metrics),
                "historicalEmployeeRecords": len(historical_employee_records),
                "historicalEmployeeNames": historical_employee_names,
                "unmatchedNewHourNames": unmatched_new_hour_names,
                "pmWorkDetailSourceNames": len(pm_work_detail_values),
                "transferHourMatchedHours": transfer_hour_notes.get("matchedHours", 0),
                "transferHourUnmatchedHours": transfer_hour_notes.get("unmatchedHours", 0),
                "rawOrderWorkbooks": raw_order_notes.get("workbooks", 0),
                "rawOrderToolsMonths": raw_order_notes.get("toolsMonths", 0),
                "rawOrderAcMonths": raw_order_notes.get("acMonths", 0),
                "rawOrderMttrMonths": raw_order_mttr_notes.get("matchedMonths", 0),
                "rawOrderMttrEmployeeMonths": raw_order_mttr_notes.get("matchedEmployeeMonths", 0),
                "acRepairHoursQualityIssueCount": ac_repair_hours_quality_notes.get("issueCount", 0),
                "acRepairHoursQualityThreshold": ac_repair_hours_quality_notes.get("threshold"),
                "rawOrderUnmappedNames": raw_order_notes.get("unmappedNames", []),
                "employeeNumberSourceNames": len(employee_numbers),
                "unmatchedPmWorkDetailNames": unmatched_pm_work_detail_names,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
