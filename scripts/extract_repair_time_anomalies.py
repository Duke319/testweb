import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
RAW_DATA_LABEL = "data/raw/employee-performance-system"
SOURCE_DIR = ROOT / RAW_DATA_LABEL / "2022-2025 pi 数据"
OUTPUT_JSON = ROOT / "data" / "repair-time-anomalies.json"

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
SHEET_NAME = "维修数据"
REPAIR_TIME_COLUMN = "维修时间(Min)"
THRESHOLD_MINUTES = 300
PREFERRED_DUPLICATE_KEYWORDS = {
    (2022, 12): "副本",
}


def decode_zip_member_name(filename):
    try:
        return filename.encode("cp437").decode("gbk")
    except Exception:
        return filename


def workbook_month_from_name(filename):
    match = re.search(r"(\d{1,2})月份维修数据", str(filename or ""))
    if not match:
        return None
    month_number = int(match.group(1))
    return month_number if 1 <= month_number <= 12 else None


def workbook_year_from_zip(filename):
    match = re.search(r"(20\d{2})", str(filename or ""))
    return int(match.group(1)) if match else None


def is_workbook_member(item):
    basename = item.filename.split("/")[-1]
    return (
        not item.is_dir()
        and not basename.startswith(("~$", "."))
        and basename.lower().endswith((".xlsx", ".xlsm"))
    )


def choose_candidate(existing, current):
    if existing is None:
        return current

    preferred_keyword = PREFERRED_DUPLICATE_KEYWORDS.get((current["year"], current["monthNumber"]))
    existing_name = existing["memberName"]
    current_name = current["memberName"]
    if preferred_keyword:
        existing_is_preferred = preferred_keyword in existing_name
        current_is_preferred = preferred_keyword in current_name
        if current_is_preferred and not existing_is_preferred:
            return current
        return existing

    if "副本" in existing_name and "副本" not in current_name:
        return current
    return existing


def collect_workbook_candidates():
    candidates = {}
    for zip_path in sorted(SOURCE_DIR.glob("*.zip")):
        year = workbook_year_from_zip(zip_path.name)
        if not year:
            continue
        with ZipFile(zip_path) as archive:
            for item in archive.infolist():
                if not is_workbook_member(item):
                    continue
                member_name = decode_zip_member_name(item.filename)
                month_number = workbook_month_from_name(member_name)
                if not month_number:
                    continue
                current = {
                    "year": year,
                    "monthNumber": month_number,
                    "month": f"{year} {MONTH_NAMES[month_number - 1]}",
                    "zipPath": zip_path,
                    "zipName": zip_path.name,
                    "memberFilename": item.filename,
                    "memberName": member_name,
                    "sourceFile": f"{RAW_DATA_LABEL}/2022-2025 pi 数据/{zip_path.name}:{member_name}",
                }
                key = (year, month_number)
                candidates[key] = choose_candidate(candidates.get(key), current)
    return [candidates[key] for key in sorted(candidates)]


def text_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def number_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, timedelta):
        return value.total_seconds() / 60
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        pass
    match = re.match(r"^(\d+):(\d{1,2})(?::(\d{1,2}))?$", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        return hours * 60 + minutes + seconds / 60
    return None


def date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 1000:
        try:
            return from_excel(value).date()
        except Exception:
            return None
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.match(r"^(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def find_column(labels, candidates):
    for candidate in candidates:
        for index, label in enumerate(labels):
            if label == candidate:
                return index
    return None


def get_cell(row, column_index):
    if column_index is None or column_index >= len(row):
        return None
    return row[column_index]


def find_header(worksheet):
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        labels = [text_value(value) for value in row]
        repair_time_column = find_column(labels, [REPAIR_TIME_COLUMN, "维修时间（Min）", "维修时间(MIN)"])
        if repair_time_column is not None and "技工" in labels:
            return row_index, labels
    return None, []


def record_id(source_file, row_number, order_no, repair_time_minutes):
    key = f"{source_file}|{row_number}|{order_no}|{repair_time_minutes:.3f}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def extract_candidate(candidate):
    records = []
    rows_scanned = 0
    with ZipFile(candidate["zipPath"]) as archive:
        workbook = load_workbook(BytesIO(archive.read(candidate["memberFilename"])), read_only=True, data_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                return records, rows_scanned
            worksheet = workbook[SHEET_NAME]
            header_row, labels = find_header(worksheet)
            if not header_row:
                return records, rows_scanned

            columns = {
                "date": find_column(labels, ["日期"]),
                "notifyDate": find_column(labels, ["通知日期"]),
                "repairResponseTime": find_column(labels, ["维修响应时间"]),
                "notifyTime": find_column(labels, ["通知时间"]),
                "repairStartTime": find_column(labels, ["维修开始时间"]),
                "repairEndTime": find_column(labels, ["维修结束时间"]),
                "downtimeMinutes": find_column(labels, ["停机时间(Min)", "停机时间（Min）"]),
                "repairTimeMinutes": find_column(labels, [REPAIR_TIME_COLUMN, "维修时间（Min）", "维修时间(MIN)"]),
                "description": find_column(labels, ["维修内容描述"]),
                "technician": find_column(labels, ["技工"]),
                "equipmentName": find_column(labels, ["设备名称"]),
                "productionLine": find_column(labels, ["生产线"]),
                "valueStream": find_column(labels, ["价值流"]),
                "problemCodeText": find_column(labels, ["问题代码文本"]),
                "reasonCodeText": find_column(labels, ["原因代码文本"]),
                "repairTimeText": find_column(labels, ["维修时间"]),
                "orderNo": find_column(labels, ["订单"]),
                "equipmentNo": find_column(labels, ["设备编号"]),
                "mainWorkCenter": find_column(labels, ["主工作中心"]),
                "systemStatus": find_column(labels, ["系统状态"]),
                "creator": find_column(labels, ["创建者"]),
                "createdAt": find_column(labels, ["创建时间"]),
                "startTime": find_column(labels, ["开始时间"]),
                "endTime": find_column(labels, ["结束时间"]),
                "faultStartDate": find_column(labels, ["故障开始日期"]),
                "downtimePeriod": find_column(labels, ["停机期间"]),
            }

            for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                repair_time_minutes = number_value(get_cell(row, columns["repairTimeMinutes"]))
                if repair_time_minutes is None or repair_time_minutes <= 0:
                    continue
                rows_scanned += 1
                if repair_time_minutes <= THRESHOLD_MINUTES:
                    continue

                row_date = date_value(get_cell(row, columns["date"]))
                year = row_date.year if row_date else candidate["year"]
                month_number = row_date.month if row_date else candidate["monthNumber"]
                month = f"{year} {MONTH_NAMES[month_number - 1]}"
                order_no = text_value(get_cell(row, columns["orderNo"]))
                source_file = candidate["sourceFile"]
                record = {
                    "id": record_id(source_file, row_number, order_no, repair_time_minutes),
                    "year": str(year),
                    "monthNumber": month_number,
                    "month": month,
                    "recordDate": row_date.isoformat() if row_date else "",
                    "employeeName": text_value(get_cell(row, columns["technician"])),
                    "technician": text_value(get_cell(row, columns["technician"])),
                    "repairTimeMinutes": round(repair_time_minutes, 1),
                    "thresholdMinutes": THRESHOLD_MINUTES,
                    "excessMinutes": round(repair_time_minutes - THRESHOLD_MINUTES, 1),
                    "repairTimeText": text_value(get_cell(row, columns["repairTimeText"])),
                    "downtimeMinutes": round(number_value(get_cell(row, columns["downtimeMinutes"])) or 0, 1),
                    "repairResponseTimeMinutes": round(number_value(get_cell(row, columns["repairResponseTime"])) or 0, 1),
                    "notifyDate": text_value(get_cell(row, columns["notifyDate"])),
                    "notifyTime": text_value(get_cell(row, columns["notifyTime"])),
                    "repairStartTime": text_value(get_cell(row, columns["repairStartTime"])),
                    "repairEndTime": text_value(get_cell(row, columns["repairEndTime"])),
                    "description": text_value(get_cell(row, columns["description"])),
                    "equipmentName": text_value(get_cell(row, columns["equipmentName"])),
                    "productionLine": text_value(get_cell(row, columns["productionLine"])),
                    "valueStream": text_value(get_cell(row, columns["valueStream"])),
                    "problemCodeText": text_value(get_cell(row, columns["problemCodeText"])),
                    "reasonCodeText": text_value(get_cell(row, columns["reasonCodeText"])),
                    "orderNo": order_no,
                    "equipmentNo": text_value(get_cell(row, columns["equipmentNo"])),
                    "mainWorkCenter": text_value(get_cell(row, columns["mainWorkCenter"])),
                    "systemStatus": text_value(get_cell(row, columns["systemStatus"])),
                    "creator": text_value(get_cell(row, columns["creator"])),
                    "createdAt": text_value(get_cell(row, columns["createdAt"])),
                    "startTime": text_value(get_cell(row, columns["startTime"])),
                    "endTime": text_value(get_cell(row, columns["endTime"])),
                    "faultStartDate": text_value(get_cell(row, columns["faultStartDate"])),
                    "downtimePeriod": text_value(get_cell(row, columns["downtimePeriod"])),
                    "businessArea": "Tools",
                    "plant": "101",
                    "department": "",
                    "workshop": "101车间",
                    "shift": "",
                    "employeeKey": "",
                    "employeeNo": "",
                    "sourceFile": source_file,
                    "sourceZip": f"{RAW_DATA_LABEL}/2022-2025 pi 数据/{candidate['zipName']}",
                    "sourceWorkbook": candidate["memberName"],
                    "sheet": SHEET_NAME,
                    "sourceRow": row_number,
                }
                records.append(record)
        finally:
            workbook.close()
    return records, rows_scanned


def build_summary(candidates, rows_scanned_by_key, records):
    yearly = defaultdict(lambda: {"rowsScanned": 0, "anomalyCount": 0, "maxRepairTimeMinutes": 0, "technicians": set()})
    monthly = defaultdict(lambda: {"rowsScanned": 0, "anomalyCount": 0, "maxRepairTimeMinutes": 0})

    for candidate in candidates:
        year = str(candidate["year"])
        month = candidate["month"]
        rows_scanned = rows_scanned_by_key[(candidate["year"], candidate["monthNumber"])]
        yearly[year]["rowsScanned"] += rows_scanned
        monthly[month]["rowsScanned"] += rows_scanned

    for record in records:
        year = record["year"]
        month = record["month"]
        repair_time = record["repairTimeMinutes"]
        yearly[year]["anomalyCount"] += 1
        yearly[year]["maxRepairTimeMinutes"] = max(yearly[year]["maxRepairTimeMinutes"], repair_time)
        if record["employeeName"]:
            yearly[year]["technicians"].add(record["employeeName"])
        monthly[month]["anomalyCount"] += 1
        monthly[month]["maxRepairTimeMinutes"] = max(monthly[month]["maxRepairTimeMinutes"], repair_time)

    yearly_rows = []
    for year, values in sorted(yearly.items()):
        rows_scanned = values["rowsScanned"]
        anomaly_count = values["anomalyCount"]
        yearly_rows.append(
            {
                "year": year,
                "rowsScanned": rows_scanned,
                "anomalyCount": anomaly_count,
                "anomalyRate": round(anomaly_count / rows_scanned, 4) if rows_scanned else 0,
                "maxRepairTimeMinutes": round(values["maxRepairTimeMinutes"], 1),
                "technicianCount": len(values["technicians"]),
            }
        )

    monthly_rows = []
    for month, values in sorted(monthly.items(), key=lambda item: (int(item[0].split()[0]), MONTH_NAMES.index(item[0].split()[1]))):
        rows_scanned = values["rowsScanned"]
        anomaly_count = values["anomalyCount"]
        monthly_rows.append(
            {
                "month": month,
                "rowsScanned": rows_scanned,
                "anomalyCount": anomaly_count,
                "anomalyRate": round(anomaly_count / rows_scanned, 4) if rows_scanned else 0,
                "maxRepairTimeMinutes": round(values["maxRepairTimeMinutes"], 1),
            }
        )

    total_rows = sum(rows_scanned_by_key.values())
    return {
        "workbookCount": len(candidates),
        "rowsScanned": total_rows,
        "anomalyCount": len(records),
        "anomalyRate": round(len(records) / total_rows, 4) if total_rows else 0,
        "yearly": yearly_rows,
        "monthly": monthly_rows,
    }


def main():
    candidates = collect_workbook_candidates()
    all_records = []
    rows_scanned_by_key = {}

    for candidate in candidates:
        records, rows_scanned = extract_candidate(candidate)
        all_records.extend(records)
        rows_scanned_by_key[(candidate["year"], candidate["monthNumber"])] = rows_scanned

    all_records.sort(key=lambda item: (item["repairTimeMinutes"], item["year"], item["monthNumber"]), reverse=True)
    summary = build_summary(candidates, rows_scanned_by_key, all_records)
    payload = {
        "source": {
            "directory": f"{RAW_DATA_LABEL}/2022-2025 pi 数据",
            "sheet": SHEET_NAME,
            "repairTimeColumn": REPAIR_TIME_COLUMN,
            "thresholdMinutes": THRESHOLD_MINUTES,
            "scope": "zip files only",
        },
        "summary": summary,
        "records": all_records,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(OUTPUT_JSON),
                "workbooks": summary["workbookCount"],
                "rowsScanned": summary["rowsScanned"],
                "anomalies": summary["anomalyCount"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
