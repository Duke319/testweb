import json
import math
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT_DIR = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT_DIR / "data" / "raw" / "employee-performance-system" / "TEF3 Near Miss.xlsx"
PERFORMANCE_FILE = ROOT_DIR / "data" / "worker-performance-monthly.json"
CERTIFICATES_FILE = ROOT_DIR / "data" / "employee-certificates.json"
OUTPUT_FILE = ROOT_DIR / "data" / "near-miss-records.json"
MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def normalize_name(value):
    text = clean_text(value)
    text = re.sub(r"^(Mr\.|Ms\.)\s*", "", text, flags=re.IGNORECASE)
    text = text.split("/")[0].strip()
    match = re.match(r"^TEF3[（(](.+?)[）)]$", text, flags=re.IGNORECASE)
    if match:
        text = match.group(1).strip()
    return text


def romanized_key(value):
    text = clean_text(value)
    text = re.sub(r"^(Mr\.|Ms\.)\s*", "", text, flags=re.IGNORECASE)
    if "/" in text:
        text = text.split("/")[-1]
    return re.sub(r"[^a-z]", "", text.lower())


def clean_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value).date().isoformat()
    parsed = datetime.fromisoformat(str(value).strip())
    return parsed.date().isoformat()


def month_parts(date_text):
    if not date_text:
        return "", "", 0, 0
    parsed = datetime.fromisoformat(date_text)
    year = str(parsed.year)
    month_number = parsed.month
    return year, f"{year} {MONTH_NAMES[month_number - 1]}", month_number, int(year) * 100 + month_number


def parse_month_label(label):
    text = clean_text(label)
    parts = text.split()
    if len(parts) != 2 or parts[1] not in MONTH_NAMES:
        return 0
    return int(parts[0]) * 100 + MONTH_NAMES.index(parts[1]) + 1


def read_performance_records():
    if not PERFORMANCE_FILE.exists():
        return []
    payload = json.loads(PERFORMANCE_FILE.read_text(encoding="utf-8"))
    return [
        record
        for record in payload.get("records", [])
        if record.get("month") and not record.get("isTotal") and record.get("employeeName") != "Total"
    ]


def read_certificate_aliases():
    if not CERTIFICATES_FILE.exists():
        return {}
    payload = json.loads(CERTIFICATES_FILE.read_text(encoding="utf-8"))
    aliases = {}
    for employee in payload.get("employees", []):
        employee_name = normalize_name(employee.get("employeeName"))
        if not employee_name:
            continue
        aliases[normalize_name(employee_name)] = employee_name
        romanized = romanized_key(employee.get("sourceName"))
        if romanized:
            aliases[romanized] = employee_name
    aliases.update(
        {
            "wangyizhi": "王一志",
            "hebiao": "何彪",
            "liuweibin": "刘卫兵",
            "liuweibing": "刘卫兵",
            "liutao": "刘涛",
            "lumengyun": "鲁孟云",
            "jiangweiqiang": "蒋卫强",
        }
    )
    return aliases


def resolve_reporter_name(value, aliases):
    name = normalize_name(value)
    if name in aliases:
        return aliases[name]
    romanized = romanized_key(name)
    return aliases.get(romanized, name)


def build_lookup(records):
    by_name = defaultdict(list)
    for record in records:
        employee_name = normalize_name(record.get("employeeName"))
        if not employee_name:
            continue
        indexed = {**record, "_month_index": parse_month_label(record.get("month"))}
        by_name[employee_name].append(indexed)

    for key in by_name:
        by_name[key].sort(key=lambda item: item["_month_index"])
    return by_name


def closest_record(records, month_index):
    if not records:
        return None
    same_month = [record for record in records if record["_month_index"] == month_index]
    if same_month:
        return same_month[-1]
    before = [record for record in records if record["_month_index"] <= month_index]
    if before:
        return before[-1]
    return records[-1]


def parse_plant(area, line):
    text = f"{clean_text(area)} {clean_text(line)}"
    match = re.search(r"\b(101|103|104)\b", text)
    return match.group(1) if match else ""


def build_record(row, row_number, by_name, aliases):
    reporter_raw = clean_text(row.get("汇报人员i姓名"))
    reporter_name = resolve_reporter_name(reporter_raw, aliases)
    occurred_date = clean_date(row.get("事故发生日期") or row.get("事故整改日期"))
    year, month, month_number, month_index = month_parts(occurred_date)
    matched = closest_record(by_name.get(reporter_name, []), month_index)
    plant = clean_text(matched.get("plant")) if matched else parse_plant(row.get("区域"), row.get("产线"))
    business_area = clean_text(matched.get("businessArea")) if matched else ("AC" if plant in {"103", "104"} else "Tools" if plant == "101" else "")
    task_id = clean_text(row.get("任务ID"))
    source_id = clean_text(row.get("序号")) or task_id or str(row_number)
    project_id = f"NEAR-MISS-{source_id}"

    return {
        "id": project_id,
        "sourceFile": str(SOURCE_FILE.relative_to(ROOT_DIR)),
        "sourceSheet": "TEF3 near miss",
        "sourceRow": row_number,
        "projectId": project_id,
        "projectTitle": clean_text(row.get("事故描述")),
        "projectType": "Near Miss",
        "improvementType": "near_miss",
        "operatorName": reporter_raw,
        "employeeNo": clean_text(matched.get("employeeNo")) if matched else "",
        "employeeName": clean_text(matched.get("employeeName")) if matched else reporter_name,
        "employeeKey": clean_text(matched.get("employeeKey")) if matched else reporter_name,
        "matchedBy": "employeeName" if matched else "",
        "matchedPerformanceMonth": clean_text(matched.get("month")) if matched else "",
        "businessArea": business_area,
        "plant": plant,
        "department": clean_text(matched.get("department")) if matched else clean_text(row.get("Department2") or row.get("事故上报的部门")),
        "workshop": clean_text(matched.get("workshop")) if matched else (f"{plant}车间" if plant else ""),
        "shift": clean_text(matched.get("shift")) if matched else "",
        "sourceDepartment": clean_text(row.get("Department")),
        "lineArea": clean_text(row.get("产线")),
        "station": clean_text(row.get("区域")),
        "createdDate": occurred_date,
        "incidentDate": occurred_date,
        "closedDate": clean_date(row.get("事故整改日期")),
        "approved": True,
        "approvalStep": "closed" if clean_text(row.get("是否关闭")) == "是" else "open",
        "year": year,
        "month": month,
        "monthNumber": month_number,
        "quantity": 1,
        "benefitAmount": 0,
        "taskId": task_id,
        "reportLink": clean_text(row.get("事故报告链接")),
        "correctiveAction": clean_text(row.get("事故措施")),
        "isClosed": clean_text(row.get("是否关闭")),
        "closeOwner": clean_text(row.get("问题关闭人员")),
    }


def read_near_miss_rows():
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    sheet = workbook["TEF3 near miss"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    records = []
    skipped_types = defaultdict(int)
    by_name = build_lookup(read_performance_records())
    aliases = read_certificate_aliases()

    for offset, values in enumerate(rows, start=2):
        if not any(values):
            continue
        row = dict(zip(headers, values))
        near_miss_type = clean_text(row.get("type")).lower()
        if near_miss_type != "near miss":
            skipped_types[clean_text(row.get("type")) or "blank"] += 1
            continue
        records.append(build_record(row, offset, by_name, aliases))

    return records, skipped_types


def build_payload():
    records, skipped_types = read_near_miss_rows()
    return {
        "source": {
            "type": "near_miss_records",
            "rawFile": str(SOURCE_FILE.relative_to(ROOT_DIR)),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sheet": "TEF3 near miss",
            "metricRule": "Rows with type = Near Miss count as near_miss; non-Near Miss rows are excluded",
        },
        "summary": {
            "rowCount": len(records) + sum(skipped_types.values()),
            "nearMissCount": len(records),
            "skippedTypes": dict(skipped_types),
            "matchedEmployeeCount": sum(1 for record in records if record["matchedBy"]),
        },
        "records": records,
    }


def main():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Near miss workbook not found: {SOURCE_FILE}")
    payload = build_payload()
    OUTPUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
